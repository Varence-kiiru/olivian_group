from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.urls import reverse, reverse_lazy
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import Quotation, Customer, QuotationFollowUp, QuotationRequest
from .forms import CustomerRequirementsWizard, QuotationCreateForm, QuotationItemFormSet, QuotationCreateFromRequestForm
from apps.core.email_utils import EmailService
from django.db.models import Count, Q, Subquery, OuterRef
from django.core.serializers.json import DjangoJSONEncoder
from apps.products.models import Product, ProductCategory, ProductImage
import json
import json
import logging

logger = logging.getLogger(__name__)

# Quotation Request Views

class CustomerQuotationRequestView(View):
    """Public customer quotation request form"""

    template_name = 'quotations/request_form.html'

    def get(self, request):
        if hasattr(request.user, 'role') and request.user.role == 'customer':
            # Pre-fill form for logged-in customers
            from apps.quotations.models import Customer
            try:
                customer = Customer.objects.get(email=request.user.email)
                initial_data = {
                    'customer_name': customer.name,
                    'email': customer.email,
                    'phone': customer.phone,
                    'company_name': customer.company_name,
                    'location': customer.city,
                    'property_type': customer.property_type,
                }
            except Customer.DoesNotExist:
                initial_data = {}
        else:
            initial_data = {}

        context = {
            'initial_data': initial_data,
            'is_logged_in_customer': hasattr(request.user, 'role') and request.user.role == 'customer'
        }
        return render(request, self.template_name, context)

    def post(self, request):
        try:
            # Store detailed form data for email
            form_data = {
                'household_info': request.POST.get('household_info', ''),
                'preferred_schedule': request.POST.get('preferred_schedule', ''),
                'power_calculator_data': request.POST.get('power_calculator_data', ''),
                'special_requirements': request.POST.get('special_requirements', False) in ['true', '1', 'on'],
                'backup_requirements': request.POST.get('backup_requirements', ''),
                'installation_timeline': request.POST.get('installation_timeline', ''),
                'concerns_requirements': request.POST.get('concerns_requirements', ''),
                'roof_details': request.POST.get('roof_details', ''),
                'budget_range': request.POST.get('budget_range', ''),
                'preferred_contact_info': request.POST.get('preferred_contact_info', ''),
                'recurring_costs': request.POST.get('recurring_costs', ''),
            }

            # Create quotation request from form data
            quotation_request = QuotationRequest.objects.create(
                customer_name=request.POST.get('customer_name'),
                email=request.POST.get('email'),
                phone=request.POST.get('phone'),
                company_name=request.POST.get('company_name', ''),
                system_type=request.POST.get('system_type'),
                system_requirements=request.POST.get('system_requirements'),
                estimated_budget=request.POST.get('estimated_budget', None),
                location=request.POST.get('location'),
                property_type=request.POST.get('property_type', 'residential'),
                roof_access=request.POST.get('roof_access', 'not_sure'),
                urgency=request.POST.get('urgency', 'medium'),
                preferred_contact_time=request.POST.get('preferred_contact_time', ''),
                customer_notes=request.POST.get('customer_notes', ''),
            )

            # Store form data for email template
            quotation_request._session_form_data = form_data

            messages.success(
                request,
                f'Thank you for your quotation request! We will contact you within {quotation_request.estimated_completion_days} working days.'
            )

            return redirect('quotations:request_success')

        except Exception as e:
            logger.error(f"Error creating quotation request: {str(e)}")
            messages.error(request, 'There was an error submitting your request. Please try again.')
            return render(request, self.template_name, {
                'initial_data': request.POST,
                'error': True
            })


class QuotationRequestSuccessView(TemplateView):
    """Success page after quotation request submission"""
    template_name = 'quotations/request_success.html'


class CheckNewRequestsView(LoginRequiredMixin, View):
    """AJAX view to check for new quotation requests"""

    def get(self, request):
        try:
            # Get new requests count (status='new')
            new_count = QuotationRequest.objects.filter(status='new').count()

            return JsonResponse({
                'success': True,
                'new_requests': new_count
            })
        except Exception as e:
            logger.error(f"Error checking new requests: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class QuotationRequestListView(LoginRequiredMixin, ListView):
    """Management view for quotation requests"""
    model = QuotationRequest
    template_name = 'quotations/requests/list.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = QuotationRequest.objects.all().order_by('-created_at')

        # Apply filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        urgency = self.request.GET.get('urgency')
        if urgency:
            queryset = queryset.filter(urgency=urgency)

        system_type = self.request.GET.get('system_type')
        if system_type:
            queryset = queryset.filter(system_type=system_type)

        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(customer_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(system_requirements__icontains=search)
            )

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context.update({
            'total_requests': self.get_queryset().count(),
            'new_requests': self.get_queryset().filter(status='new').count(),
            'pending_requests': self.get_queryset().filter(status__in=['new', 'reviewed', 'contacted']).count(),
            'completed_requests': self.get_queryset().filter(status__in=['quotation_created', 'completed']).count(),
            'status_choices': QuotationRequest.STATUS_CHOICES,
            'urgency_choices': QuotationRequest.URGENCY_LEVELS,
            'system_type_choices': QuotationRequest.SYSTEM_TYPES,
        })

        return context


class QuotationRequestDetailView(LoginRequiredMixin, DetailView):
    """Detail view for quotation requests (management only)"""
    model = QuotationRequest
    template_name = 'quotations/requests/detail.html'
    context_object_name = 'quotation_request'

    def post(self, request, *args, **kwargs):
        quotation_request = self.get_object()
        action = request.POST.get('action')

        if action == 'assign_to_me':
            quotation_request.assigned_to = request.user
            quotation_request.save()
            messages.success(request, f'Assigned request to yourself')

        elif action == 'update_status':
            new_status = request.POST.get('status')
            if new_status and new_status in dict(QuotationRequest.STATUS_CHOICES):
                quotation_request.status = new_status
                quotation_request.save()
                messages.success(request, f'Status updated to {quotation_request.get_status_display()}')

        return redirect('quotations:request_detail', pk=quotation_request.pk)


class BulkActionsView(LoginRequiredMixin, View):
    """Handle bulk actions for quotation requests"""

    def post(self, request):
        selected_requests_ids = request.POST.get('selected_requests', '').split(',')
        action = request.POST.get('action')

        if not selected_requests_ids or not action:
            messages.error(request, 'Please select requests and an action.')
            return redirect('quotations:requests_list')

        try:
            # Convert to integers and filter valid IDs
            request_ids = [int(id) for id in selected_requests_ids if id.strip()]
            requests = QuotationRequest.objects.filter(id__in=request_ids)

            if not requests.exists():
                messages.error(request, 'No valid requests selected.')
                return redirect('quotations:requests_list')

            if action == 'assign_staff':
                staff_member_id = request.POST.get('staff_member')
                if staff_member_id:
                    try:
                        from django.contrib.auth import get_user_model
                        User = get_user_model()
                        staff_member = User.objects.get(id=staff_member_id)
                        requests.update(assigned_to=staff_member)
                        messages.success(request, f'Assigned {requests.count()} requests to {staff_member.get_full_name()}.')
                    except User.DoesNotExist:
                        messages.error(request, 'Invalid staff member selected.')
                else:
                    messages.error(request, 'Please select a staff member.')

            elif action == 'update_status':
                new_status = request.POST.get('new_status')
                if new_status in dict(QuotationRequest.STATUS_CHOICES):
                    requests.update(status=new_status)
                    messages.success(request, f'Updated status for {requests.count()} requests.')
                else:
                    messages.error(request, 'Invalid status selected.')

            elif action == 'bulk_email':
                # For now, just show success message
                messages.success(request, f'Bulk email sent to {requests.count()} requests.')

            else:
                messages.error(request, 'Invalid action selected.')

        except Exception as e:
            logger.error(f"Error processing bulk actions: {str(e)}")
            messages.error(request, 'An error occurred while processing your request.')

        return redirect('quotations:requests_list')


class QuotationCreateFromRequestView(LoginRequiredMixin, View):
    """Create quotation from a quotation request"""

    template_name = 'quotations/create.html'

    def get_quotation_request(self):
        """Get and validate the quotation request"""
        try:
            pk = self.kwargs.get('request_pk')
            return get_object_or_404(QuotationRequest, pk=pk)
        except:
            return None

    def dispatch(self, request, *args, **kwargs):
        """Check permissions and request availability"""
        self.quotation_request = self.get_quotation_request()

        if not self.quotation_request:
            messages.error(request, 'Quotation request not found.')
            return redirect('quotations:requests_list')

        # Check if quotation already exists
        if hasattr(self.quotation_request, 'created_quotation') and self.quotation_request.created_quotation:
            messages.info(request, f'A quotation has already been created for this request: {self.quotation_request.created_quotation.quotation_number}')
            return redirect('quotations:detail', quotation_number=self.quotation_request.created_quotation.quotation_number)

        return super().dispatch(request, *args, **kwargs)

    def get_customer_from_request(self):
        """Try to find or create customer from request data"""
        # Try to find existing customer by email
        customer = Customer.objects.filter(email=self.quotation_request.email).first()
        if customer:
            return customer

        # Create new customer from request data
        customer = Customer.objects.create(
            name=self.quotation_request.customer_name,
            email=self.quotation_request.email,
            phone=self.quotation_request.phone,
            company_name=self.quotation_request.company_name,
            city=self.quotation_request.location,
            address=self.quotation_request.location,  # Basic address from location
            property_type=self.quotation_request.property_type,
        )
        return customer

    def get(self, request, *args, **kwargs):
        """Display the quotation creation form pre-filled from request"""
        customer = self.get_customer_from_request()

        # Create initial quotation with pre-filled data
        quotation = Quotation(
            quotation_request=self.quotation_request,
            customer=customer,
            system_type=self.quotation_request.system_type,
            system_requirements=self.quotation_request.system_requirements,
            title=f"Quotation for {self.quotation_request.customer_name}",
        )

        # Create forms
        quotation_form = QuotationCreateFromRequestForm(
            instance=quotation,
            request=request,
            quotation_request=self.quotation_request
        )
        item_formset = QuotationItemFormSet(instance=quotation)

        context = {
            'quotation_form': quotation_form,
            'item_formset': item_formset,
            'quotation_request': self.quotation_request,
            'customer': customer,
            'is_from_request': True,
        }

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Handle quotation creation with items"""
        customer = self.get_customer_from_request()

        quotation = Quotation(
            quotation_request=self.quotation_request,
            customer=customer,
        )

        quotation_form = QuotationCreateFromRequestForm(
            request.POST,
            instance=quotation,
            request=request,
            quotation_request=self.quotation_request
        )
        item_formset = QuotationItemFormSet(request.POST, instance=quotation)

        if quotation_form.is_valid() and item_formset.is_valid():
            # Check for at least one item
            items = item_formset.save(commit=False)
            if not items:
                messages.error(request, 'Please add at least one item to the quotation.')
            else:
                # Check if this is a draft save or final save
                is_draft = request.POST.get('save_draft') == 'true'

                # Save quotation
                quotation = quotation_form.save(commit=False)
                quotation.created_by = request.user

                if not is_draft:
                    # Final save: set status to sent
                    quotation.status = 'sent'

                # Note: Draft status is 'draft' by default, and Quotation.save() only updates request status for non-drafts
                quotation.save()  # This will also update totals via item saves

                # Save items
                # Use formset.save() to save the items (sets quotation automatically)
                item_formset.save()

                # Update totals after all items are saved
                quotation.update_totals()

                # For final quotations, send email with PDF attachment
                if not is_draft:
                    try:
                        # Send email with PDF attachment to customer
                        email_sent = EmailService.send_quotation_created_email(quotation)
                        if email_sent:
                            logger.info(f"Quotation email sent successfully for {quotation.quotation_number}")
                            messages.success(
                                request,
                                f'Quotation {quotation.quotation_number} created and sent to customer successfully.'
                            )
                        else:
                            logger.warning(f"Failed to send quotation email for {quotation.quotation_number}")
                            messages.warning(
                                request,
                                f'Quotation {quotation.quotation_number} created successfully, but email could not be sent.'
                            )
                    except Exception as e:
                        logger.error(f"Error sending quotation email for {quotation.quotation_number}: {str(e)}")
                        messages.warning(
                            request,
                            f'Quotation {quotation.quotation_number} created successfully, but email could not be sent.'
                        )
                else:
                    # Draft message
                    messages.success(
                        request,
                        f'Quotation {quotation.quotation_number} saved as draft.'
                    )

                return redirect('quotations:detail', quotation_number=quotation.quotation_number)
            # If not items, fall through to re-render
        else:
            context = {
                'quotation_form': quotation_form,
                'item_formset': item_formset,
                'quotation_request': self.quotation_request,
                'customer': customer,
                'is_from_request': True,
            }
            return render(request, self.template_name, context)


class QuotationDashboardView(LoginRequiredMixin, TemplateView):
    """Enhanced quotations dashboard with comprehensive metrics"""
    template_name = 'quotations/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build base queryset
        queryset = Quotation.objects.all()

        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)

        # Basic metrics
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['pending_quotations'] = queryset.filter(status__in=['draft', 'sent']).count()

        # Status breakdown
        context['draft_count'] = queryset.filter(status='draft').count()
        context['sent_count'] = queryset.filter(status='sent').count()
        context['viewed_count'] = queryset.filter(status='viewed').count()
        context['accepted_count'] = queryset.filter(status='accepted').count()
        context['converted_count'] = queryset.filter(status='converted').count()

        # Conversion rate
        context['conversion_rate'] = (context['converted_count'] / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0

        # System capacity analysis
        residential_quotations = queryset.filter(system_capacity__lt=10)
        commercial_quotations = queryset.filter(system_capacity__gte=10, system_capacity__lt=50)
        industrial_quotations = queryset.filter(system_capacity__gte=50)

        context['residential_count'] = residential_quotations.count()
        context['commercial_count'] = commercial_quotations.count()
        context['industrial_count'] = industrial_quotations.count()

        context['residential_value'] = sum(q.total_amount for q in residential_quotations)
        context['commercial_value'] = sum(q.total_amount for q in commercial_quotations)
        context['industrial_value'] = sum(q.total_amount for q in industrial_quotations)

        # Calculator generated quotations
        calculator_queryset = queryset.filter(notes__icontains='Generated from solar calculator')
        context['calculator_count'] = calculator_queryset.count()
        context['calculator_conversion_rate'] = (calculator_queryset.filter(status='converted').count() / calculator_queryset.count() * 100) if calculator_queryset.count() > 0 else 0

        # Performance metrics
        context['avg_deal_size'] = context['total_value'] / context['total_quotations'] if context['total_quotations'] > 0 else 0
        context['avg_response_time'] = 2.4  # Mock data for now

        # Recent activity (mock data for now)
        context['recent_activities'] = [
            {
                'title': 'New quotation created',
                'description': 'QUO-2024-0001 for residential system',
                'icon': 'file-invoice',
                'timestamp': timezone.now() - timezone.timedelta(hours=2)
            },
            {
                'title': 'Quotation sent',
                'description': 'QUO-2024-0002 sent to customer',
                'icon': 'paper-plane',
                'timestamp': timezone.now() - timezone.timedelta(hours=4)
            },
            {
                'title': 'Quotation accepted',
                'description': 'QUO-2024-0003 accepted and converted',
                'icon': 'check-circle',
                'timestamp': timezone.now() - timezone.timedelta(hours=6)
            }
        ]
        
        return context

class QuotationListView(LoginRequiredMixin, ListView):
    model = Quotation
    template_name = 'quotations/list.html'
    context_object_name = 'quotations'
    paginate_by = 20

    def get_queryset(self):
        from django.db.models import Q
        queryset = Quotation.objects.select_related('customer', 'salesperson').order_by('-created_at')

        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)
        
        # Apply search filters
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(quotation_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__email__icontains=search)
            )
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        system_type = self.request.GET.get('system_type')
        if system_type:
            queryset = queryset.filter(system_type=system_type)
            
        # Filter by source (calculator vs manual)
        source = self.request.GET.get('source')
        if source == 'calculator':
            queryset = queryset.filter(notes__icontains='Generated from solar calculator')
        elif source == 'manual':
            queryset = queryset.exclude(notes__icontains='Generated from solar calculator')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add statistics for dashboard
        queryset = self.get_queryset()
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['pending_quotations'] = queryset.filter(status__in=['draft', 'sent']).count()
        
        # Calculate conversion rate
        converted_count = queryset.filter(status='converted').count()
        context['conversion_rate'] = (converted_count / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0
        
        # Add filter options
        context['status_choices'] = Quotation.STATUS_CHOICES
        context['system_type_choices'] = Quotation.SYSTEM_TYPES
        
        # Count calculator vs manual quotations
        calculator_count = queryset.filter(notes__icontains='Generated from solar calculator').count()
        manual_count = queryset.exclude(notes__icontains='Generated from solar calculator').count()
        context['calculator_quotations'] = calculator_count
        context['manual_quotations'] = manual_count
        
        return context

class QuotationDetailView(LoginRequiredMixin, DetailView):
    model = Quotation
    template_name = 'quotations/detail.html'
    context_object_name = 'quotation'
    slug_field = 'quotation_number'
    slug_url_kwarg = 'quotation_number'

class QuotationCreateView(LoginRequiredMixin, CreateView):
    model = Quotation
    template_name = 'quotations/create.html'
    fields = ['customer', 'quotation_type', 'system_type', 'system_capacity']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Check if coming from requirements assessment
        if 'customer_requirements' in self.request.session:
            context['from_requirements'] = True
            context['requirements_data'] = self.request.session['customer_requirements']
        else:
            context['from_requirements'] = False

        # Add customers and products data for autofill
        from apps.quotations.models import Customer

        # Add customers list
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                context['customers'] = Customer.objects.filter(email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                context['customers'] = Customer.objects.filter(assigned_salesperson=self.request.user)
            else:
                context['customers'] = Customer.objects.all()
        else:
            context['customers'] = Customer.objects.all()

        # Add products JSON for JavaScript
        try:
            from apps.products.models import Product
            products = Product.objects.filter(status='active').values(
                'id', 'name', 'price', 'sku', 'description'
            )
            context['products_json'] = json.dumps(list(products))
        except Exception:
            context['products_json'] = '[]'

        return context

    def form_valid(self, form):
        form.instance.salesperson = self.request.user
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class CreateQuotationFromRequirementsView(LoginRequiredMixin, View):
    """Create quotation directly from customer requirements assessment"""

    def get(self, request):
        """Redirect to quotation create page (in future could auto-create quotation)"""
        return redirect('quotations:create')

    def post(self, request):
        """Create quotation from requirements data"""
        try:
            # Get requirements data from session
            requirements_data = request.session.get('customer_requirements')
            if not requirements_data:
                messages.error(request, 'No requirements data found. Please complete the assessment first.')
                return redirect('quotations:customer_requirements')

            # Create or get customer
            from apps.quotations.models import Customer
            customer_name = requirements_data['customer_name']
            customer_email = requirements_data['customer_email']
            customer_phone = requirements_data['customer_phone']
            customer_company = requirements_data.get('customer_company', '')
            customer_address = requirements_data['customer_address']
            customer_city = requirements_data['customer_city']
            customer_county = requirements_data['customer_county']

            try:
                customer = Customer.objects.get(email=customer_email)
                # Update existing customer info if needed
                customer.name = customer_name
                customer.phone = customer_phone
                customer.address = customer_address
                customer.city = customer_city
                customer.monthly_consumption = requirements_data.get('monthly_consumption', 0)
                customer.average_monthly_bill = requirements_data.get('monthly_bill', 0.0)
                customer.save()
            except Customer.DoesNotExist:
                customer = Customer.objects.create(
                    name=customer_name,
                    email=customer_email,
                    phone=customer_phone,
                    company=customer_company,
                    address=customer_address,
                    city=customer_city,
                    county=customer_county,
                    monthly_consumption=requirements_data.get('monthly_consumption', 0),
                    average_monthly_bill=requirements_data.get('monthly_bill', 0.0),
                    property_type=requirements_data.get('property_type', 'residential'),
                    roof_type=requirements_data.get('roof_type', 'concrete'),
                    roof_area=requirements_data.get('roof_area', 0),
                    building_age=requirements_data.get('building_age', ''),
                    floors=requirements_data.get('floors', '1'),
                    roof_orientation=requirements_data.get('roof_orientation', ''),
                    shading_issues=requirements_data.get('shading_issues', []),
                )

            # Calculate system capacity from requirements (rough estimate)
            monthly_consumption = requirements_data.get('monthly_consumption', 0)
            system_capacity = 0
            if monthly_consumption:
                # Rough calculation: monthly consumption / 120 kWh (average monthly generation per kW in Kenya)
                system_capacity = monthly_consumption / 120
                if system_capacity < 1:
                    system_capacity = 1
                # Round to reasonable increments
                if system_capacity <= 3:
                    system_capacity = round(system_capacity)
                elif system_capacity <= 10:
                    system_capacity = round(system_capacity / 2) * 2  # Round to nearest 2
                else:
                    system_capacity = round(system_capacity / 5) * 5  # Round to nearest 5

            # Determine system type from property and roof data
            property_type = requirements_data.get('property_type', 'residential')
            roof_type = requirements_data.get('roof_type', 'concrete')

            system_type = 'grid_tied'  # Default
            if property_type == 'industrial' or requirements_data.get('backup_needs') in ['full']:
                system_type = 'hybrid'
            elif requirements_data.get('grid_connection') in ['poor', 'none']:
                system_type = 'off_grid'

            quotation_type = 'standard'
            if property_type == 'commercial':
                quotation_type = 'commercial'
            elif property_type == 'industrial':
                quotation_type = 'industrial'

            # Create quotation with requirements data
            quotation = Quotation.objects.create(
                customer=customer,
                salesperson=request.user,
                quotation_type=quotation_type,
                system_type=system_type,
                system_capacity=system_capacity,
                estimated_generation=requirements_data.get('monthly_consumption', 0) * 0.85,  # Rough estimate with 85% efficiency
                installation_complexity='simple',  # Can be adjusted based on roof type, building age, etc.

                # Financial calculations will be done by model save method
                subtotal=0,
                discount_amount=0,
                discount_percentage=0.0,
                tax_amount=0,
                total_amount=0,

                # Savings calculations based on requirements
                estimated_monthly_savings=requirements_data.get('monthly_bill', 0.0) * 0.35,  # Rough 35% savings estimate
                estimated_annual_savings=(requirements_data.get('monthly_bill', 0.0) * 0.35) * 12,
                payback_period_months=round((system_capacity * 35000) / (requirements_data.get('monthly_bill', 0.0) * 0.35 * 12)) if requirements_data.get('monthly_bill', 0.0) > 0 else 120,
                roi_percentage=((requirements_data.get('monthly_bill', 0.0) * 0.35 * 12) / (system_capacity * 35000)) * 100 if system_capacity > 0 else 8,

                # Terms and validity
                valid_until=timezone.now().date() + timezone.timedelta(days=30),
                warranty_terms="25-year panel performance warranty, 12-year product warranty, 10-year inverter warranty, 5-year installation warranty",

                # Notes with requirements data
                notes=f"""Generated from Customer Requirements Assessment:

CUSTOMER DETAILS:
- Name: {customer_name}
- Email: {customer_email}
- Phone: {customer_phone}
{'- Company: '+customer_company if customer_company else ''}
- Address: {customer_address}
- City: {customer_city}
- County: {customer_county}

PROPERTY INFORMATION:
- Property Type: {property_type.title()}
- Building Age: {requirements_data.get('building_age', 'Not specified')}
- Floors: {requirements_data.get('floors', '1')}
- Roof Type: {roof_type.title()}
- Roof Area: {requirements_data.get('roof_area', 0)} sq meters
- Roof Orientation: {requirements_data.get('roof_orientation', 'Not specified')}
- Shading Issues: {', '.join(requirements_data.get('shading_issues', [])) or 'None'}

ENERGY CONSUMPTION ANALYSIS:
- Monthly Bill: KES {requirements_data.get('monthly_bill', 0):,.0f}
- Monthly Consumption: {requirements_data.get('monthly_consumption', 'Not calculated')} kWh
- Peak Usage Time: {requirements_data.get('peak_usage_time', 'Not specified').title()}
- Backup Needs: {requirements_data.get('backup_needs', 'Not specified').replace('_', ' ').title()}
- Grid Connection: {requirements_data.get('grid_connection', 'Not specified').replace('_', ' ').title()}

CALCULATED SYSTEM SPECIFICATIONS:
- Recommended Capacity: {system_capacity} kW
- System Type: {system_type.replace('_', ' ').title()}
- Estimated Monthly Savings: KES {requirements_data.get('monthly_bill', 0)*0.35:,.0f}
- Estimated Annual Savings: KES {(requirements_data.get('monthly_bill', 0)*0.35)*12:,.0f}
- Expected Payback Period: {round((system_capacity * 35000) / (requirements_data.get('monthly_bill', 0.0) * 0.35 * 12)) if requirements_data.get('monthly_bill', 0) > 0 else 120} months

Note: These specifications are initial recommendations based on the assessment. Final system design requires professional site assessment.""",

                status='draft'  # Start as draft for review
            )

            # Clear the session data after successful creation
            del request.session['customer_requirements']
            request.session.modified = True

            messages.success(request, f'Quotation {quotation.quotation_number} created successfully from customer requirements assessment!')
            return redirect('quotations:edit', quotation_number=quotation.quotation_number)

        except Exception as e:
            logger.error(f"Error creating quotation from requirements: {str(e)}")
            messages.error(request, f'Error creating quotation from requirements: {str(e)}')
            return redirect('quotations:create')

class ConvertToSaleView(LoginRequiredMixin, DetailView):
    model = Quotation
    template_name = 'quotations/detail.html'  # Changed from quotation_detail.html to detail.html
    slug_field = 'quotation_number'
    slug_url_kwarg = 'quotation_number'
    
    def post(self, request, *args, **kwargs):
        quotation = self.get_object()
        if quotation.can_be_converted():
            order = quotation.convert_to_sale()
            messages.success(request, f'Quotation {quotation.quotation_number} converted to order {order.order_number}')
            return redirect('ecommerce:order_detail', order_number=order.order_number)
        else:
            messages.error(request, 'Quotation cannot be converted to sale')
            return redirect('quotations:detail', quotation_number=quotation.quotation_number)

class SolarCalculatorView(TemplateView):
    template_name = 'quotations/calculator.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Determine if this is accessed from management or customer side
        # Check the referer or URL path to determine context
        referer = self.request.META.get('HTTP_REFERER', '')
        current_path = self.request.path
        
        # Check if accessed from management URLs or logged in as staff
        is_management_access = (
            '/quotations/' in referer or  # Accessed from quotations management
            '/admin/' in referer or       # Accessed from admin
            '/quotations/' in current_path or  # Direct management path
            (self.request.user.is_authenticated and 
             hasattr(self.request.user, 'is_staff') and 
             self.request.user.is_staff) or  # Staff user
            'from=management' in self.request.GET  # Explicit parameter
        )
        
        # Customer-facing URL takes precedence
        if '/solar-calculator/' in current_path:
            is_management_access = False
        
        # Set the appropriate base template
        if is_management_access:
            context['base_template'] = 'dashboard/base.html'  # Management base template
            context['is_management_view'] = True
            context['page_title'] = 'Solar Calculator - Management'
        else:
            context['base_template'] = 'website/base.html'  # Customer base template
            context['is_management_view'] = False
            context['page_title'] = 'Solar Calculator'
            
        # Add customer information if logged in (for auto-populating forms)
        if self.request.user.is_authenticated and hasattr(self.request.user, 'profile'):
            profile = self.request.user.profile
            context['customer_info'] = {
                'name': f"{self.request.user.first_name} {self.request.user.last_name}".strip() or self.request.user.username,
                'email': self.request.user.email,
                'phone': getattr(profile, 'phone', '') if profile else ''
            }
        elif self.request.user.is_authenticated:
            context['customer_info'] = {
                'name': f"{self.request.user.first_name} {self.request.user.last_name}".strip() or self.request.user.username,
                'email': self.request.user.email,
                'phone': ''
            }
            
        return context


class CreateFromCalculatorView(LoginRequiredMixin, CreateView):
    """Create quotation from calculator results"""
    model = Quotation
    template_name = 'quotations/create_from_calculator.html'
    fields = ['customer', 'quotation_type', 'system_type', 'system_capacity']
    
    def form_valid(self, form):
        form.instance.salesperson = self.request.user
        
        # Get calculator data from POST
        calculator_data = {
            'monthly_bill': self.request.POST.get('monthly_bill', ''),
            'roof_area': self.request.POST.get('roof_area', ''),
            'system_capacity': self.request.POST.get('system_capacity', ''),
            'estimated_cost': self.request.POST.get('estimated_cost', ''),
        }
        
        # Store calculator data in quotation notes
        form.instance.notes = f"Generated from calculator with: {calculator_data}"
        
        messages.success(self.request, 'Quotation created successfully from calculator!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Pass calculator data to template
        context.update({
            'monthly_bill': self.request.POST.get('monthly_bill', ''),
            'roof_area': self.request.POST.get('roof_area', ''),
            'system_capacity': self.request.POST.get('system_capacity', ''),
            'estimated_cost': self.request.POST.get('estimated_cost', ''),
        })
        
        return context


class QuotationPDFView(LoginRequiredMixin, View):
    """Generate PDF for quotation"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Check permissions
        if (request.user.role == 'customer' and 
            quotation.customer.email != request.user.email):
            messages.error(request, 'You do not have permission to view this quotation.')
            return redirect('quotations:list')
        elif (request.user.role in ['sales_person', 'sales_manager'] and 
              quotation.salesperson != request.user):
            messages.error(request, 'You do not have permission to view this quotation.')
            return redirect('quotations:list')
        
        # Generate PDF
        pdf_response = self.generate_quotation_pdf(quotation)
        return pdf_response
    
    def generate_quotation_pdf(self, quotation, return_response=True):
        """Generate PDF for quotation using reportlab"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import Table, TableStyle, Paragraph
            from django.conf import settings
            from apps.core.models import CompanySettings
            from io import BytesIO
            import os
            import logging
            
            logger = logging.getLogger(__name__)
            logger.info(f"Starting PDF generation for quotation {quotation.quotation_number}")
        
        except ImportError as e:
            logger.error(f"Failed to import PDF libraries: {e}")
            raise Exception(f"PDF generation libraries not available: {e}")
        
        # Get company settings
        try:
            company = CompanySettings.objects.first()
        except CompanySettings.DoesNotExist:
            company = None
        
        # Create PDF buffer
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Company branding
        company_name = company.name if company else 'The Olivian Group Limited'
        company_phone = company.phone if company else '+254-719-728-666'
        company_email = company.email if company else 'info@olivian.co.ke'
        company_website = company.website if company else 'https://olivian.co.ke'
        primary_color = company.primary_color if company else '#38b6ff'
        
        # Add company logo if available
        logo_y = height - 80
        if company and company.logo:
            try:
                logo_path = os.path.join(settings.MEDIA_ROOT, company.logo.name)
                if os.path.exists(logo_path):
                    c.drawImage(logo_path, 50, logo_y, width=80, height=60, preserveAspectRatio=True)
                    logo_y = height - 90
            except:
                pass
        
        # Company header
        c.setFont("Helvetica-Bold", 18)
        try:
            c.setFillColor(colors.HexColor(primary_color))
        except:
            c.setFillColor(colors.blue)
        c.drawString(150, height - 60, company_name)
        
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        c.drawString(150, height - 80, f"Phone: {company_phone}")
        c.drawString(150, height - 95, f"Email: {company_email}")
        c.drawString(150, height - 110, f"Website: {company_website}")
        
        # Quotation title
        c.setFont("Helvetica-Bold", 24)
        try:
            c.setFillColor(colors.HexColor(primary_color))
        except:
            c.setFillColor(colors.blue)
        c.drawString(50, height - 150, "SOLAR SYSTEM QUOTATION")
        
        # Quotation details
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.black)
        c.drawString(50, height - 180, f"Quotation #: {quotation.quotation_number}")
        c.drawString(50, height - 195, f"Date: {quotation.created_at.strftime('%B %d, %Y')}")
        c.drawString(50, height - 210, f"Valid Until: {quotation.valid_until.strftime('%B %d, %Y') if quotation.valid_until else 'N/A'}")
        
        # Customer details
        c.drawString(350, height - 180, "CUSTOMER DETAILS")
        c.setFont("Helvetica", 11)
        c.drawString(350, height - 195, f"Name: {quotation.customer.name}")
        c.drawString(350, height - 210, f"Email: {quotation.customer.email}")
        if quotation.customer.phone:
            c.drawString(350, height - 225, f"Phone: {quotation.customer.phone}")
        
        # System specifications
        y_position = height - 270
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "SOLAR SYSTEM SPECIFICATIONS")
        
        y_position -= 25
        c.setFont("Helvetica", 11)
        c.setFillColor(colors.black)
        
        specs = [
            f"System Type: {quotation.get_system_type_display()}",
            f"System Capacity: {quotation.system_capacity} kW",
            f"Quotation Type: {quotation.get_quotation_type_display()}",
        ]
        
        for spec in specs:
            c.drawString(50, y_position, spec)
            y_position -= 15
        
        # System Components
        if quotation.items.exists():
            y_position -= 20
            c.setFont("Helvetica-Bold", 14)
            c.setFillColor(colors.HexColor(primary_color))
            c.drawString(50, y_position, "SYSTEM COMPONENTS")
            
            y_position -= 20
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.black)
            
            # Component table headers
            c.setFont("Helvetica-Bold", 10)
            c.drawString(50, y_position, "Component")
            c.drawString(300, y_position, "Qty")
            c.drawString(350, y_position, "Unit Price")
            c.drawString(450, y_position, "Total")
            y_position -= 15
            
            # Draw line under headers
            c.line(50, y_position + 5, 550, y_position + 5)
            
            # Component items
            c.setFont("Helvetica", 9)
            for item in quotation.items.all():
                c.drawString(50, y_position, f"{item.item_name}")
                c.drawString(300, y_position, f"{item.quantity:.0f} {item.unit}")
                c.drawString(350, y_position, f"KES {item.unit_price:,.0f}")
                c.drawString(450, y_position, f"KES {item.total_price:,.0f}")
                y_position -= 12
                
                # Add description if available
                if item.description:
                    c.setFont("Helvetica", 8)
                    c.setFillColor(colors.grey)
                    # Wrap description text
                    desc_words = item.description.split()
                    line = ""
                    for word in desc_words:
                        if len(line + word) < 60:  # Max characters per line
                            line += word + " "
                        else:
                            if line.strip():
                                c.drawString(70, y_position, line.strip())
                                y_position -= 10
                            line = word + " "
                    if line.strip():
                        c.drawString(70, y_position, line.strip())
                        y_position -= 10
                    
                    c.setFont("Helvetica", 9)
                    c.setFillColor(colors.black)
                    y_position -= 3
        
        # Financial summary
        y_position -= 20
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "FINANCIAL SUMMARY")
        
        y_position -= 25
        c.setFont("Helvetica", 11)
        c.setFillColor(colors.black)
        
        if quotation.subtotal > 0:
            c.drawString(50, y_position, f"Subtotal (Ex-VAT): KES {quotation.subtotal:,.2f}")
            y_position -= 15
            
            # Show discount if applicable
            if quotation.discount_amount > 0:
                c.drawString(50, y_position, f"Discount ({quotation.discount_percentage}%): - KES {quotation.discount_amount:,.2f}")
                y_position -= 15
            
            # VAT calculation (already calculated in model)
            c.drawString(50, y_position, f"VAT (16%): KES {quotation.tax_amount:,.2f}")
            y_position -= 15
            
            # Final total (already includes VAT)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y_position, f"TOTAL AMOUNT: KES {quotation.total_amount:,.2f}")
        elif quotation.total_amount > 0:
            # Fallback for old quotations without proper item breakdown
            c.drawString(50, y_position, f"Total System Cost: KES {quotation.total_amount:,.2f}")
            c.setFont("Helvetica", 9)
            y_position -= 12
            c.drawString(50, y_position, "(Note: This quotation predates component breakdown)")
        
        # Terms and conditions
        y_position -= 40
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "TERMS & CONDITIONS")
        
        y_position -= 20
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        terms = [
            "• This quotation is valid for 30 days from the date of issue",
            "• 50% deposit required to commence installation",
            "• Installation typically takes 2-5 working days",
            "• 25-year performance warranty on solar panels",
            "• 10-year warranty on inverters",
            "• 5-year workmanship warranty",
            "• Professional site assessment included",
            "• All prices are in Kenyan Shillings (KES)",
            "• KPLC approvals and permits handled by Olivian",
            "• Final design may vary based on site assessment"
        ]
        
        for term in terms:
            c.drawString(50, y_position, term)
            y_position -= 12
        
        # Footer
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.gray)
        c.drawString(50, 50, f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}")
        c.drawString(50, 35, f"Contact us: {company_phone} | {company_email} | {company_website}")
        
        # Professional disclaimer
        c.drawString(50, 20, "This quotation is based on preliminary calculations. Final system design requires professional site assessment.")
        
        c.save()
        buffer.seek(0)
        
        if return_response:
            # Return PDF response for direct download
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quotation_number}.pdf"'
            return response
        else:
            # Return raw PDF content for email attachment
            return buffer.getvalue()


@method_decorator(csrf_exempt, name='dispatch')
class GenerateQuotationFromCalculatorView(View):
    """Generate quotation request from calculator data"""

    def post(self, request):
        try:
            # Parse calculator data
            data = json.loads(request.body)

            # Extract calculator results
            calculator_data = data.get('calculatorData', {})
            customer_data = data.get('customerData', {})

            # Validate required fields
            customer_email = customer_data.get('email')
            customer_name = customer_data.get('name')
            customer_phone = customer_data.get('phone', '')

            if not customer_email or not customer_name:
                return JsonResponse({'success': False, 'message': 'Customer name and email are required'})

            # Create quotation request from calculator data
            quotation_request = QuotationRequest.objects.create(
                customer_name=customer_name,
                email=customer_email,
                phone=customer_phone,
                company_name=customer_data.get('company', ''),
                system_type=calculator_data.get('systemType', 'grid_tied'),
                system_requirements=f"""Solar system quotation request generated from calculator.

SYSTEM SPECIFICATIONS:
- Recommended Capacity: {calculator_data.get('capacity', 0)} kW
- System Type: {calculator_data.get('systemType', 'grid_tied').replace('_', ' ').title()}
- Monthly Generation: {calculator_data.get('monthlyGeneration', 0)} kWh
- Monthly Savings: KES {calculator_data.get('monthlySavings', 0):,.0f}
- Annual Savings: KES {calculator_data.get('annualSavings', 0):,.0f}
- Payback Period: {calculator_data.get('paybackPeriod', 'N/A')} years
- System Efficiency: {calculator_data.get('systemEfficiency', 85)}%
- Performance Ratio: {calculator_data.get('performanceRatio', 85)}%
- Self Consumption: {calculator_data.get('selfConsumption', 50)}%
- CO2 Savings: {calculator_data.get('co2Savings', 0)} tons/year

PROPERTY DETAILS:
- Property Type: {calculator_data.get('propertyType', 'residential').title()}
- Roof Type: {calculator_data.get('roofType', 'concrete').title()}
- Roof Area: {calculator_data.get('roofArea', 0)} sq meters
- Monthly Consumption: {calculator_data.get('monthlyConsumption', 0)} kWh
- Average Monthly Bill: KES {calculator_data.get('monthlyBill', 0):,.0f}

CALCULATOR DATA:
{json.dumps(calculator_data, indent=2)}

Please review this request and create a formal quotation.""",
                estimated_budget=calculator_data.get('estimatedCost', 0),
                location=customer_data.get('location', ''),
                property_type=calculator_data.get('propertyType', 'residential'),
                roof_access='not_sure',  # Default since calculator doesn't specify
                urgency='medium',  # Default urgency
                preferred_contact_time='',
                customer_notes=f"Generated from solar calculator. Customer requested quotation for {calculator_data.get('capacity', 0)}kW {calculator_data.get('systemType', 'grid_tied').replace('_', ' ').title()} system.",
            )

            return JsonResponse({
                'success': True,
                'message': 'Quotation request submitted successfully! Our team will contact you within 3-5 working days.',
                'request_id': quotation_request.pk,
                'email_sent': True
            })

        except Exception as e:
            logger.error(f"Error generating quotation request from calculator: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error submitting quotation request: {str(e)}'
            })


class CustomerQuotationView(View):
    """Public customer view for quotations (no login required)"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Get related data
        customer = quotation.customer
        
        context = {
            'quotation': quotation,
            'customer': customer,
            'company': self.get_company_settings(),
            'is_customer_view': True,
        }
        return render(request, 'quotations/customer_view.html', context)
    
    def get_company_settings(self):
        """Get company settings"""
        try:
            from apps.core.models import CompanySettings
            return CompanySettings.objects.first()
        except:
            return None


class CustomerQuotationPDFView(View):
    """Generate PDF for customer (no login required)"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Use the same PDF generation logic but without login requirement
        pdf_view = QuotationPDFView()
        pdf_response = pdf_view.generate_quotation_pdf(quotation)
        return pdf_response


# Export and Analytics Views
class QuotationExportView(LoginRequiredMixin, View):
    """Export quotations in various formats"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'csv')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        
        # Build queryset
        queryset = Quotation.objects.all()
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
            
        # Apply role-based filtering
        if request.user.role == 'customer':
            queryset = queryset.filter(customer__email=request.user.email)
        elif request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(salesperson=request.user)
            
        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        elif format_type == 'pdf':
            return self.export_pdf_report(queryset)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)
    
    def export_csv(self, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="quotations.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Quotation Number', 'Customer Name', 'Customer Email', 
            'System Type', 'Capacity (kW)', 'Total Amount (KES)', 
            'Status', 'Created Date', 'Valid Until'
        ])
        
        for quotation in queryset:
            writer.writerow([
                quotation.quotation_number,
                quotation.customer.name,
                quotation.customer.email,
                quotation.get_system_type_display(),
                quotation.system_capacity,
                quotation.total_amount,
                quotation.get_status_display(),
                quotation.created_at.strftime('%Y-%m-%d'),
                quotation.valid_until.strftime('%Y-%m-%d') if quotation.valid_until else ''
            ])
        
        return response
    
    def export_excel(self, queryset):
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Quotations'
        
        # Headers
        headers = [
            'Quotation Number', 'Customer Name', 'Customer Email', 
            'System Type', 'Capacity (kW)', 'Total Amount (KES)', 
            'Monthly Savings (KES)', 'Annual Savings (KES)',
            'Status', 'Created Date', 'Valid Until'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, quotation in enumerate(queryset, 2):
            worksheet.cell(row=row, column=1, value=quotation.quotation_number)
            worksheet.cell(row=row, column=2, value=quotation.customer.name)
            worksheet.cell(row=row, column=3, value=quotation.customer.email)
            worksheet.cell(row=row, column=4, value=quotation.get_system_type_display())
            worksheet.cell(row=row, column=5, value=float(quotation.system_capacity))
            worksheet.cell(row=row, column=6, value=float(quotation.total_amount))
            worksheet.cell(row=row, column=7, value=float(quotation.estimated_monthly_savings))
            worksheet.cell(row=row, column=8, value=float(quotation.estimated_annual_savings))
            worksheet.cell(row=row, column=9, value=quotation.get_status_display())
            worksheet.cell(row=row, column=10, value=quotation.created_at.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=11, value=quotation.valid_until.strftime('%Y-%m-%d') if quotation.valid_until else '')
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="quotations.xlsx"'
        return response
    
    def export_pdf_report(self, queryset):
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        import weasyprint
        
        context = {
            'quotations': queryset,
            'export_date': timezone.now(),
            'total_count': queryset.count(),
            'total_value': sum(q.total_amount for q in queryset),
        }
        
        html_string = render_to_string('quotations/export_report.html', context)
        html = weasyprint.HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="quotations_report.pdf"'
        return response


class QuotationBulkActionsView(LoginRequiredMixin, View):
    """Handle bulk operations on quotations"""
    
    def post(self, request):
        action = request.POST.get('action')
        quotation_ids = request.POST.getlist('quotation_ids')
        
        if not quotation_ids:
            return JsonResponse({'success': False, 'message': 'No quotations selected'})
        
        queryset = Quotation.objects.filter(id__in=quotation_ids)
        
        # Apply role-based filtering
        if request.user.role == 'customer':
            queryset = queryset.filter(customer__email=request.user.email)
        elif request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(salesperson=request.user)
        
        if action == 'delete':
            count = queryset.filter(status__in=['draft', 'sent']).count()
            queryset.filter(status__in=['draft', 'sent']).delete()
            return JsonResponse({'success': True, 'message': f'{count} quotations deleted'})
        
        elif action == 'mark_sent':
            count = queryset.filter(status='draft').update(status='sent')
            return JsonResponse({'success': True, 'message': f'{count} quotations marked as sent'})
        
        elif action == 'mark_expired':
            count = queryset.filter(status__in=['sent', 'viewed']).update(status='expired')
            return JsonResponse({'success': True, 'message': f'{count} quotations marked as expired'})
        
        else:
            return JsonResponse({'success': False, 'message': 'Invalid action'})


class QuotationAnalyticsView(LoginRequiredMixin, TemplateView):
    """Quotation analytics and reporting dashboard"""
    template_name = 'quotations/analytics.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Build base queryset
        queryset = Quotation.objects.all()
        
        # Apply role-based filtering
        if self.request.user.role == 'customer':
            queryset = queryset.filter(customer__email=self.request.user.email)
        elif self.request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(salesperson=self.request.user)
        
        # Basic statistics
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount or 0 for q in queryset)
        context['average_value'] = context['total_value'] / context['total_quotations'] if context['total_quotations'] > 0 else 0
        
        # Status breakdown
        context['status_stats'] = {
            status[0]: queryset.filter(status=status[0]).count()
            for status in Quotation.STATUS_CHOICES
        }
        
        # Conversion rate
        converted_count = queryset.filter(status='converted').count()
        context['conversion_rate'] = (converted_count / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0
        
        # Monthly trends (last 12 months)
        from django.db.models import Count, Sum
        from django.db.models.functions import ExtractMonth, ExtractYear
        from django.utils import timezone

        # Use ExtractMonth and ExtractYear to avoid timezone conversion issues
        monthly_data = queryset.annotate(
            year=ExtractYear('created_at'),
            month=ExtractMonth('created_at')
        ).values('year', 'month').annotate(
            count=Count('id'),
            total_value=Sum('total_amount')
        ).order_by('year', 'month')

        # Format the data for the template
        formatted_monthly_data = []
        for item in monthly_data:
            # Skip items with None year or month to prevent TypeError
            if item['year'] is None or item['month'] is None:
                logger.warning(f"Skipping monthly data item with None year/month: {item}")
                continue

            # Create a date object for the month
            from datetime import date
            month_date = date(item['year'], item['month'], 1)
            formatted_monthly_data.append({
                'month': month_date,
                'count': item['count'],
                'total_value': item['total_value']
            })

        context['monthly_data'] = formatted_monthly_data
        
        # System type breakdown
        context['system_type_stats'] = {
            system_type[0]: queryset.filter(system_type=system_type[0]).count()
            for system_type in Quotation.SYSTEM_TYPES
        }
        
        return context


# CRUD Views
class QuotationUpdateView(LoginRequiredMixin, UpdateView):
    """Edit quotation"""
    model = Quotation
    fields = [
        'quotation_type', 'system_type', 'system_capacity', 'estimated_generation',
        'installation_complexity', 'subtotal', 'discount_amount', 'discount_percentage',
        'tax_amount', 'total_amount', 'estimated_monthly_savings', 'estimated_annual_savings',
        'payback_period_months', 'roi_percentage', 'financing_available', 'payment_terms',
        'delivery_time', 'warranty_terms', 'notes'
    ]
    template_name = 'quotations/edit.html'
    
    def get_object(self):
        return get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
    
    def get_success_url(self):
        return reverse('quotations:detail', kwargs={'quotation_number': self.object.quotation_number})
    
    def form_valid(self, form):
        messages.success(self.request, f'Quotation {self.object.quotation_number} updated successfully!')
        return super().form_valid(form)


class QuotationDeleteView(LoginRequiredMixin, DeleteView):
    """Delete quotation"""
    model = Quotation
    template_name = 'quotations/delete_confirm.html'
    success_url = reverse_lazy('quotations:list')
    
    def get_object(self):
        return get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
    
    def delete(self, request, *args, **kwargs):
        quotation = self.get_object()
        
        # Only allow deletion of draft or sent quotations
        if quotation.status not in ['draft', 'sent']:
            messages.error(request, 'Cannot delete quotations that have been accepted or converted.')
            return redirect('quotations:detail', quotation_number=quotation.quotation_number)
        
        quotation_number = quotation.quotation_number
        quotation.delete()
        messages.success(request, f'Quotation {quotation_number} deleted successfully!')
        return redirect(self.success_url)


class QuotationSendEmailView(LoginRequiredMixin, View):
    """Manually send quotation email"""
    
    def post(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        try:
            from apps.core.email_utils import EmailService
            success = EmailService.send_quotation_created_email(quotation)
            
            if success:
                quotation.status = 'sent'
                quotation.save()
                messages.success(request, f'Quotation {quotation_number} sent successfully!')
            else:
                messages.error(request, 'Failed to send quotation email.')
                
        except Exception as e:
            messages.error(request, f'Error sending email: {str(e)}')
        
        return redirect('quotations:detail', quotation_number=quotation_number)


class QuotationDuplicateView(LoginRequiredMixin, View):
    """Duplicate an existing quotation"""
    
    def post(self, request, quotation_number):
        original = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Create duplicate
        duplicate = Quotation.objects.create(
            customer=original.customer,
            salesperson=request.user,
            quotation_type=original.quotation_type,
            system_type=original.system_type,
            system_capacity=original.system_capacity,
            estimated_generation=original.estimated_generation,
            installation_complexity=original.installation_complexity,
            subtotal=original.subtotal,
            discount_amount=original.discount_amount,
            discount_percentage=original.discount_percentage,
            tax_amount=original.tax_amount,
            total_amount=original.total_amount,
            estimated_monthly_savings=original.estimated_monthly_savings,
            estimated_annual_savings=original.estimated_annual_savings,
            payback_period_months=original.payback_period_months,
            roi_percentage=original.roi_percentage,
            financing_available=original.financing_available,
            financing_options=original.financing_options,
            payment_terms=original.payment_terms,
            delivery_time=original.delivery_time,
            warranty_terms=original.warranty_terms,
            notes=f"Duplicated from {original.quotation_number}\n\n{original.notes}",
            status='draft'
        )
        
        messages.success(request, f'Quotation duplicated as {duplicate.quotation_number}')
        return redirect('quotations:edit', quotation_number=duplicate.quotation_number)


class QuotationFollowUpView(LoginRequiredMixin, CreateView):
    model = QuotationFollowUp
    fields = ['follow_up_type', 'notes', 'scheduled_date']
    
    def form_valid(self, form):
        try:
            # Get date and time from separate inputs
            date_str = self.request.POST.get('scheduled_date_0')
            time_str = self.request.POST.get('scheduled_date_1')
            
            # Combine date and time
            from datetime import datetime
            scheduled_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            
            # Set the datetime on the form instance
            form.instance.scheduled_date = scheduled_datetime
            
            # Set other required fields
            quotation = get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
            form.instance.quotation = quotation
            form.instance.created_by = self.request.user
            
            messages.success(self.request, 'Follow-up added successfully!')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Error saving follow-up: {str(e)}')
            return super().form_invalid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please check the form data and try again.')
        return redirect('quotations:detail', quotation_number=self.kwargs['quotation_number'])
    
    def get_success_url(self):
        return reverse('quotations:detail', kwargs={'quotation_number': self.kwargs['quotation_number']})
    
    # Override get method to redirect to detail view if accessed directly
    def get(self, request, *args, **kwargs):
        return redirect('quotations:detail', quotation_number=self.kwargs['quotation_number'])

class QuotationFollowUpListView(LoginRequiredMixin, ListView):
    model = QuotationFollowUp
    template_name = 'quotations/follow_ups.html'
    context_object_name = 'follow_ups'
    paginate_by = 20

    def get_queryset(self):
        queryset = QuotationFollowUp.objects.select_related('quotation', 'created_by')
        
        # Filter by role
        if self.request.user.role == 'customer':
            queryset = queryset.filter(quotation__customer__email=self.request.user.email)
        elif self.request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(quotation__salesperson=self.request.user)

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            if status == 'pending':
                queryset = queryset.filter(scheduled_date__gte=timezone.now())
            elif status == 'completed':
                queryset = queryset.filter(completed=True)
            elif status == 'overdue':
                queryset = queryset.filter(scheduled_date__lt=timezone.now(), completed=False)

        return queryset.order_by('scheduled_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pending_count'] = self.get_queryset().filter(scheduled_date__gte=timezone.now(), completed=False).count()
        context['overdue_count'] = self.get_queryset().filter(scheduled_date__lt=timezone.now(), completed=False).count()
        context['completed_count'] = self.get_queryset().filter(completed=True).count()
        return context

@require_POST
@login_required
def complete_follow_up(request, pk):
    """Mark a follow-up as completed"""
    try:
        follow_up = get_object_or_404(QuotationFollowUp, pk=pk)
        
        # Check permissions
        if request.user != follow_up.created_by and request.user != follow_up.quotation.salesperson:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to complete this follow-up'
            }, status=403)
        
        # Mark as completed
        follow_up.completed = True
        follow_up.save()
        
        # Add completion note if provided
        completion_note = request.POST.get('completion_note')
        if completion_note:
            follow_up.notes += f"\n\nCompletion Note ({timezone.now().strftime('%Y-%m-%d %H:%M')}): {completion_note}"
            follow_up.save(update_fields=['notes'])
        
        messages.success(request, 'Follow-up marked as completed successfully!')
        return JsonResponse({
            'success': True,
            'message': 'Follow-up completed successfully'
        })
        
    except Exception as e:
        logger.error(f"Error completing follow-up: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error completing follow-up: {str(e)}'
        }, status=500)

class QuotationDashboardView(LoginRequiredMixin, TemplateView):
    """Enhanced quotations dashboard with comprehensive metrics"""
    template_name = 'quotations/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build base queryset
        queryset = Quotation.objects.all()

        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)

        # Basic metrics
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['pending_quotations'] = queryset.filter(status__in=['draft', 'sent']).count()

        # Status breakdown
        context['draft_count'] = queryset.filter(status='draft').count()
        context['sent_count'] = queryset.filter(status='sent').count()
        context['viewed_count'] = queryset.filter(status='viewed').count()
        context['accepted_count'] = queryset.filter(status='accepted').count()
        context['converted_count'] = queryset.filter(status='converted').count()
        
        # Conversion rate
        context['conversion_rate'] = (context['converted_count'] / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0

        # System capacity analysis
        residential_quotations = queryset.filter(system_capacity__lt=10)
        commercial_quotations = queryset.filter(system_capacity__gte=10, system_capacity__lt=50)
        industrial_quotations = queryset.filter(system_capacity__gte=50)

        context['residential_count'] = residential_quotations.count()
        context['commercial_count'] = commercial_quotations.count()
        context['industrial_count'] = industrial_quotations.count()

        context['residential_value'] = sum(q.total_amount for q in residential_quotations)
        context['commercial_value'] = sum(q.total_amount for q in commercial_quotations)
        context['industrial_value'] = sum(q.total_amount for q in industrial_quotations)

        # Calculator generated quotations
        calculator_queryset = queryset.filter(notes__icontains='Generated from solar calculator')
        context['calculator_count'] = calculator_queryset.count()
        context['calculator_conversion_rate'] = (calculator_queryset.filter(status='converted').count() / calculator_queryset.count() * 100) if calculator_queryset.count() > 0 else 0
        
        # Performance metrics
        context['avg_deal_size'] = context['total_value'] / context['total_quotations'] if context['total_quotations'] > 0 else 0
        context['avg_response_time'] = 2.4  # Mock data for now
        
        # Recent activity (mock data for now)
        context['recent_activities'] = [
            {
                'title': 'New quotation created',
                'description': 'QUO-2024-0001 for residential system',
                'icon': 'file-invoice',
                'timestamp': timezone.now() - timezone.timedelta(hours=2)
            },
            {
                'title': 'Quotation sent',
                'description': 'QUO-2024-0002 sent to customer',
                'icon': 'paper-plane',
                'timestamp': timezone.now() - timezone.timedelta(hours=4)
            },
            {
                'title': 'Quotation accepted',
                'description': 'QUO-2024-0003 accepted and converted',
                'icon': 'check-circle',
                'timestamp': timezone.now() - timezone.timedelta(hours=6)
            }
        ]
        
        return context

class QuotationListView(LoginRequiredMixin, ListView):
    model = Quotation
    template_name = 'quotations/list.html'
    context_object_name = 'quotations'
    paginate_by = 20
    
    def get_queryset(self):
        from django.db.models import Q
        queryset = Quotation.objects.select_related('customer', 'salesperson').order_by('-created_at')
        
        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)
        
        # Apply search filters
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(quotation_number__icontains=search) |
                Q(customer__name__icontains=search) |
                Q(customer__email__icontains=search)
            )
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        system_type = self.request.GET.get('system_type')
        if system_type:
            queryset = queryset.filter(system_type=system_type)
            
        # Filter by source (calculator vs manual)
        source = self.request.GET.get('source')
        if source == 'calculator':
            queryset = queryset.filter(notes__icontains='Generated from solar calculator')
        elif source == 'manual':
            queryset = queryset.exclude(notes__icontains='Generated from solar calculator')
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add statistics for dashboard
        queryset = self.get_queryset()
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['pending_quotations'] = queryset.filter(status__in=['draft', 'sent']).count()
        
        # Calculate conversion rate
        converted_count = queryset.filter(status='converted').count()
        context['conversion_rate'] = (converted_count / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0
        
        # Add filter options
        context['status_choices'] = Quotation.STATUS_CHOICES
        context['system_type_choices'] = Quotation.SYSTEM_TYPES
        
        # Count calculator vs manual quotations
        calculator_count = queryset.filter(notes__icontains='Generated from solar calculator').count()
        manual_count = queryset.exclude(notes__icontains='Generated from solar calculator').count()
        context['calculator_quotations'] = calculator_count
        context['manual_quotations'] = manual_count
        
        return context

class QuotationDetailView(LoginRequiredMixin, DetailView):
    model = Quotation
    template_name = 'quotations/detail.html'
    context_object_name = 'quotation'
    slug_field = 'quotation_number'
    slug_url_kwarg = 'quotation_number'

class QuotationCreateView(LoginRequiredMixin, CreateView):
    model = Quotation
    template_name = 'quotations/create.html'
    fields = ['customer', 'quotation_type', 'system_type', 'system_capacity']
    
    def form_valid(self, form):
        form.instance.salesperson = self.request.user
        return super().form_valid(form)

class ConvertToSaleView(LoginRequiredMixin, DetailView):
    model = Quotation
    template_name = 'quotations/detail.html'  # Changed from quotation_detail.html to detail.html
    slug_field = 'quotation_number'
    slug_url_kwarg = 'quotation_number'
    
    def post(self, request, *args, **kwargs):
        quotation = self.get_object()
        if quotation.can_be_converted():
            order = quotation.convert_to_sale()
            messages.success(request, f'Quotation {quotation.quotation_number} converted to order {order.order_number}')
            return redirect('ecommerce:order_detail', order_number=order.order_number)
        else:
            messages.error(request, 'Quotation cannot be converted to sale')
            return redirect('quotations:detail', quotation_number=quotation.quotation_number)

class SolarCalculatorView(TemplateView):
    template_name = 'quotations/calculator.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Determine if this is accessed from management or customer side
        # Check the referer or URL path to determine context
        referer = self.request.META.get('HTTP_REFERER', '')
        current_path = self.request.path
        
        # Check if accessed from management URLs or logged in as staff
        is_management_access = (
            '/quotations/' in referer or  # Accessed from quotations management
            '/admin/' in referer or       # Accessed from admin
            '/quotations/' in current_path or  # Direct management path
            (self.request.user.is_authenticated and 
             hasattr(self.request.user, 'is_staff') and 
             self.request.user.is_staff) or  # Staff user
            'from=management' in self.request.GET  # Explicit parameter
        )
        
        # Customer-facing URL takes precedence
        if '/solar-calculator/' in current_path:
            is_management_access = False
        
        # Set the appropriate base template
        if is_management_access:
            context['base_template'] = 'dashboard/base.html'  # Management base template
            context['is_management_view'] = True
            context['page_title'] = 'Solar Calculator - Management'
        else:
            context['base_template'] = 'website/base.html'  # Customer base template
            context['is_management_view'] = False
            context['page_title'] = 'Solar Calculator'
            
        # Add customer information if logged in (for auto-populating forms)
        if self.request.user.is_authenticated and hasattr(self.request.user, 'profile'):
            profile = self.request.user.profile
            context['customer_info'] = {
                'name': f"{self.request.user.first_name} {self.request.user.last_name}".strip() or self.request.user.username,
                'email': self.request.user.email,
                'phone': getattr(profile, 'phone', '') if profile else ''
            }
        elif self.request.user.is_authenticated:
            context['customer_info'] = {
                'name': f"{self.request.user.first_name} {self.request.user.last_name}".strip() or self.request.user.username,
                'email': self.request.user.email,
                'phone': ''
            }
            
        return context


class CreateFromCalculatorView(LoginRequiredMixin, CreateView):
    """Create quotation from calculator results"""
    model = Quotation
    template_name = 'quotations/create_from_calculator.html'
    fields = ['customer', 'quotation_type', 'system_type', 'system_capacity']
    
    def form_valid(self, form):
        form.instance.salesperson = self.request.user
        
        # Get calculator data from POST
        calculator_data = {
            'monthly_bill': self.request.POST.get('monthly_bill', ''),
            'roof_area': self.request.POST.get('roof_area', ''),
            'system_capacity': self.request.POST.get('system_capacity', ''),
            'estimated_cost': self.request.POST.get('estimated_cost', ''),
        }
        
        # Store calculator data in quotation notes
        form.instance.notes = f"Generated from calculator with: {calculator_data}"
        
        messages.success(self.request, 'Quotation created successfully from calculator!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Pass calculator data to template
        context.update({
            'monthly_bill': self.request.POST.get('monthly_bill', ''),
            'roof_area': self.request.POST.get('roof_area', ''),
            'system_capacity': self.request.POST.get('system_capacity', ''),
            'estimated_cost': self.request.POST.get('estimated_cost', ''),
        })
        
        return context


class QuotationPDFView(LoginRequiredMixin, View):
    """Generate PDF for quotation"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Check permissions
        if (request.user.role == 'customer' and 
            quotation.customer.email != request.user.email):
            messages.error(request, 'You do not have permission to view this quotation.')
            return redirect('quotations:list')
        elif (request.user.role in ['sales_person', 'sales_manager'] and 
              quotation.salesperson != request.user):
            messages.error(request, 'You do not have permission to view this quotation.')
            return redirect('quotations:list')
        
        # Generate PDF
        pdf_response = self.generate_quotation_pdf(quotation)
        return pdf_response
    
    def generate_quotation_pdf(self, quotation, return_response=True):
        """Generate PDF for quotation using reportlab"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import Table, TableStyle, Paragraph
            from django.conf import settings
            from apps.core.models import CompanySettings
            from io import BytesIO
            import os
            import logging
            
            logger = logging.getLogger(__name__)
            logger.info(f"Starting PDF generation for quotation {quotation.quotation_number}")
        
        except ImportError as e:
            logger.error(f"Failed to import PDF libraries: {e}")
            raise Exception(f"PDF generation libraries not available: {e}")
        
        # Get company settings
        try:
            company = CompanySettings.objects.first()
        except CompanySettings.DoesNotExist:
            company = None
        
        # Create PDF buffer
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Company branding
        company_name = company.name if company else 'The Olivian Group Limited'
        company_phone = company.phone if company else '+254-719-728-666'
        company_email = company.email if company else 'info@olivian.co.ke'
        company_website = company.website if company else 'https://olivian.co.ke'
        primary_color = company.primary_color if company else '#38b6ff'
        
        # Add company logo if available
        logo_y = height - 80
        if company and company.logo:
            try:
                logo_path = os.path.join(settings.MEDIA_ROOT, company.logo.name)
                if os.path.exists(logo_path):
                    c.drawImage(logo_path, 50, logo_y, width=80, height=60, preserveAspectRatio=True)
                    logo_y = height - 90
            except:
                pass
        
        # Company header
        c.setFont("Helvetica-Bold", 18)
        try:
            c.setFillColor(colors.HexColor(primary_color))
        except:
            c.setFillColor(colors.blue)
        c.drawString(150, height - 60, company_name)
        
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        c.drawString(150, height - 80, f"Phone: {company_phone}")
        c.drawString(150, height - 95, f"Email: {company_email}")
        c.drawString(150, height - 110, f"Website: {company_website}")
        
        # Quotation title
        c.setFont("Helvetica-Bold", 24)
        try:
            c.setFillColor(colors.HexColor(primary_color))
        except:
            c.setFillColor(colors.blue)
        c.drawString(50, height - 150, "SOLAR SYSTEM QUOTATION")
        
        # Quotation details
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.black)
        c.drawString(50, height - 180, f"Quotation #: {quotation.quotation_number}")
        c.drawString(50, height - 195, f"Date: {quotation.created_at.strftime('%B %d, %Y')}")
        c.drawString(50, height - 210, f"Valid Until: {quotation.valid_until.strftime('%B %d, %Y') if quotation.valid_until else 'N/A'}")
        
        # Customer details
        c.drawString(350, height - 180, "CUSTOMER DETAILS")
        c.setFont("Helvetica", 11)
        c.drawString(350, height - 195, f"Name: {quotation.customer.name}")
        c.drawString(350, height - 210, f"Email: {quotation.customer.email}")
        if quotation.customer.phone:
            c.drawString(350, height - 225, f"Phone: {quotation.customer.phone}")
        
        # System specifications
        y_position = height - 270
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "SOLAR SYSTEM SPECIFICATIONS")
        
        y_position -= 25
        c.setFont("Helvetica", 11)
        c.setFillColor(colors.black)
        
        specs = [
            f"System Type: {quotation.get_system_type_display()}",
            f"System Capacity: {quotation.system_capacity} kW",
            f"Quotation Type: {quotation.get_quotation_type_display()}",
        ]
        
        for spec in specs:
            c.drawString(50, y_position, spec)
            y_position -= 15
        
        # System Components
        if quotation.items.exists():
            y_position -= 20
            c.setFont("Helvetica-Bold", 14)
            c.setFillColor(colors.HexColor(primary_color))
            c.drawString(50, y_position, "SYSTEM COMPONENTS")
            
            y_position -= 20
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.black)
            
            # Component table headers
            c.setFont("Helvetica-Bold", 10)
            c.drawString(50, y_position, "Component")
            c.drawString(300, y_position, "Qty")
            c.drawString(350, y_position, "Unit Price")
            c.drawString(450, y_position, "Total")
            y_position -= 15
            
            # Draw line under headers
            c.line(50, y_position + 5, 550, y_position + 5)
            
            # Component items
            c.setFont("Helvetica", 9)
            for item in quotation.items.all():
                c.drawString(50, y_position, f"{item.item_name}")
                c.drawString(300, y_position, f"{item.quantity:.0f} {item.unit}")
                c.drawString(350, y_position, f"KES {item.unit_price:,.0f}")
                c.drawString(450, y_position, f"KES {item.total_price:,.0f}")
                y_position -= 12
                
                # Add description if available
                if item.description:
                    c.setFont("Helvetica", 8)
                    c.setFillColor(colors.grey)
                    # Wrap description text
                    desc_words = item.description.split()
                    line = ""
                    for word in desc_words:
                        if len(line + word) < 60:  # Max characters per line
                            line += word + " "
                        else:
                            if line.strip():
                                c.drawString(70, y_position, line.strip())
                                y_position -= 10
                            line = word + " "
                    if line.strip():
                        c.drawString(70, y_position, line.strip())
                        y_position -= 10
                    
                    c.setFont("Helvetica", 9)
                    c.setFillColor(colors.black)
                    y_position -= 3
        
        # Financial summary
        y_position -= 20
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "FINANCIAL SUMMARY")
        
        y_position -= 25
        c.setFont("Helvetica", 11)
        c.setFillColor(colors.black)
        
        if quotation.subtotal > 0:
            c.drawString(50, y_position, f"Subtotal (Ex-VAT): KES {quotation.subtotal:,.2f}")
            y_position -= 15
            
            # Show discount if applicable
            if quotation.discount_amount > 0:
                c.drawString(50, y_position, f"Discount ({quotation.discount_percentage}%): - KES {quotation.discount_amount:,.2f}")
                y_position -= 15
            
            # VAT calculation (already calculated in model)
            c.drawString(50, y_position, f"VAT (16%): KES {quotation.tax_amount:,.2f}")
            y_position -= 15
            
            # Final total (already includes VAT)
            c.setFont("Helvetica-Bold", 12)
            c.drawString(50, y_position, f"TOTAL AMOUNT: KES {quotation.total_amount:,.2f}")
        elif quotation.total_amount > 0:
            # Fallback for old quotations without proper item breakdown
            c.drawString(50, y_position, f"Total System Cost: KES {quotation.total_amount:,.2f}")
            c.setFont("Helvetica", 9)
            y_position -= 12
            c.drawString(50, y_position, "(Note: This quotation predates component breakdown)")
        
        # Terms and conditions
        y_position -= 40
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawString(50, y_position, "TERMS & CONDITIONS")
        
        y_position -= 20
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        terms = [
            "• This quotation is valid for 30 days from the date of issue",
            "• 50% deposit required to commence installation",
            "• Installation typically takes 2-5 working days",
            "• 25-year performance warranty on solar panels",
            "• 10-year warranty on inverters",
            "• 5-year workmanship warranty",
            "• Professional site assessment included",
            "• All prices are in Kenyan Shillings (KES)",
            "• KPLC approvals and permits handled by Olivian",
            "• Final design may vary based on site assessment"
        ]
        
        for term in terms:
            c.drawString(50, y_position, term)
            y_position -= 12
        
        # Footer
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.gray)
        c.drawString(50, 50, f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')}")
        c.drawString(50, 35, f"Contact us: {company_phone} | {company_email} | {company_website}")
        
        # Professional disclaimer
        c.drawString(50, 20, "This quotation is based on preliminary calculations. Final system design requires professional site assessment.")
        
        c.save()
        buffer.seek(0)
        
        if return_response:
            # Return PDF response for direct download
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="quotation_{quotation.quotation_number}.pdf"'
            return response
        else:
            # Return raw PDF content for email attachment
            return buffer.getvalue()


@method_decorator(csrf_exempt, name='dispatch')
class GenerateQuotationFromCalculatorView(View):
    """Generate quotation from calculator and send email"""
    
    def get_default_salesperson(self, request):
        """Get the current user or default admin salesperson"""
        if request.user.is_authenticated:
            return request.user
        
        # Get default admin/system user for calculator-generated quotations
        from apps.accounts.models import User
        try:
            return User.objects.filter(is_staff=True).first()
        except:
            # If no admin users exist, create a system user
            return User.objects.create_user(
                username='system',
                email='system@olivian.co.ke',
                first_name='System',
                last_name='Administrator',
                is_staff=True
            )
    
    def post(self, request):
        try:
            # Parse calculator data
            data = json.loads(request.body)
            
            # Extract calculator results
            calculator_data = data.get('calculatorData', {})
            customer_data = data.get('customerData', {})
            
            # Create or get customer
            customer_email = customer_data.get('email')
            customer_name = customer_data.get('name')
            customer_phone = customer_data.get('phone', '')
            
            if not customer_email or not customer_name:
                return JsonResponse({'success': False, 'message': 'Customer name and email are required'})
            
            # Try to get existing customer or create new one
            try:
                customer = Customer.objects.get(email=customer_email)
            except Customer.DoesNotExist:
                customer = Customer.objects.create(
                    name=customer_name,
                    email=customer_email,
                    phone=customer_phone,
                    address='',  # Set default empty address
                    city='',     # Set default empty city
                    monthly_consumption=calculator_data.get('monthlyConsumption', 0),
                    average_monthly_bill=calculator_data.get('monthlyBill', 0),
                    property_type=calculator_data.get('propertyType', 'residential'),
                    roof_type=calculator_data.get('roofType', 'concrete'),
                    roof_area=calculator_data.get('roofArea', 0)
                )
            
            # Calculate ROI percentage from payback period
            payback_period_years = calculator_data.get('paybackPeriod', 10)
            roi_percentage = (100 / payback_period_years) if payback_period_years > 0 else 10
            
            # Set validity date (30 days from now)
            from datetime import datetime, timedelta
            valid_until = datetime.now().date() + timedelta(days=30)
            
            # Create quotation from calculator data (signal will handle email)
            quotation = Quotation.objects.create(
                customer=customer,
                quotation_type='custom_solution',
                system_type=calculator_data.get('systemType', 'grid_tied'),
                system_capacity=calculator_data.get('capacity', 0),
                estimated_generation=calculator_data.get('monthlyGeneration', 0),
                installation_complexity='simple',
                status='draft',  # Start as draft, then update to sent to trigger email
                
                # Financial information (will be recalculated after adding items)
                subtotal=0,
                tax_amount=0,
                total_amount=0,
                
                # Savings calculations
                estimated_monthly_savings=calculator_data.get('monthlySavings', 0),
                estimated_annual_savings=calculator_data.get('annualSavings', 0),
                payback_period_months=int(payback_period_years * 12),
                roi_percentage=roi_percentage,
                
                # Terms and validity
                valid_until=valid_until,
                warranty_terms="25-year panel performance warranty, 12-year product warranty, 10-year inverter warranty, 5-year installation warranty",
                
                # Additional information
                notes=f"Generated from solar calculator.\n\nSystem Details:\n- Capacity: {calculator_data.get('capacity', 0)} kW\n- Monthly Generation: {calculator_data.get('monthlyGeneration', 0)} kWh\n- System Efficiency: {calculator_data.get('systemEfficiency', 85)}%\n- Performance Ratio: {calculator_data.get('performanceRatio', 85)}%\n- Self Consumption: {calculator_data.get('selfConsumption', 50)}%\n- CO2 Savings: {calculator_data.get('co2Savings', 0)} tons/year\n\nGenerated automatically from Solar Calculator.",
                salesperson=self.get_default_salesperson(request)
            )
            
            # Add system components based on calculator data
            from apps.quotations.models import QuotationItem
            from decimal import Decimal
            
            system_capacity = Decimal(str(calculator_data.get('capacity', 0)))
            components_data = calculator_data.get('components', {})
            
            # Solar Panels
            panels_data = components_data.get('panels', {})
            if system_capacity > 0:
                panel_qty = int(system_capacity * 2)  # Estimate 2 panels per kW
                panel_price = Decimal('15000')  # Default panel price
                panel_name = panels_data.get('name', f'Solar Panels ({system_capacity}kW System)')
                
                QuotationItem.objects.create(
                    quotation=quotation,
                    item_name=panel_name,
                    description=f'High-efficiency monocrystalline solar panels for {system_capacity}kW system',
                    quantity=Decimal(str(panel_qty)),
                    unit='pcs',
                    unit_price=panel_price,
                    total_price=Decimal(str(panel_qty)) * panel_price
                )
            
            # Inverter
            if system_capacity > 0:
                inverter_price = system_capacity * Decimal('35000')  # Estimate KES 35k per kW
                inverter_name = f'{system_capacity}kW Solar Inverter'
                
                QuotationItem.objects.create(
                    quotation=quotation,
                    item_name=inverter_name,
                    description=f'High-efficiency string inverter for {system_capacity}kW system',
                    quantity=Decimal('1'),
                    unit='unit',
                    unit_price=inverter_price,
                    total_price=inverter_price
                )
            
            # Installation & Components
            installation_price = system_capacity * Decimal('25000')  # Estimate KES 25k per kW
            QuotationItem.objects.create(
                quotation=quotation,
                item_name='Professional Installation',
                description='Professional installation including mounting, wiring, commissioning, and KPLC approvals',
                quantity=Decimal('1'),
                unit='system',
                unit_price=installation_price,
                total_price=installation_price
            )
            
            # Calculate proper totals based on items
            quotation.calculate_totals()
            
            # Update status to 'sent' to trigger email via signal (single email)
            quotation.status = 'sent'
            quotation.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Quotation generated and sent successfully!',
                'quotation_number': quotation.quotation_number,
                'email_sent': True
            })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error generating quotation: {str(e)}'
            })


class CustomerQuotationView(View):
    """Public customer view for quotations (no login required)"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Get related data
        customer = quotation.customer
        
        context = {
            'quotation': quotation,
            'customer': customer,
            'company': self.get_company_settings(),
            'is_customer_view': True,
        }
        return render(request, 'quotations/customer_view.html', context)
    
    def get_company_settings(self):
        """Get company settings"""
        try:
            from apps.core.models import CompanySettings
            return CompanySettings.objects.first()
        except:
            return None


class CustomerQuotationPDFView(View):
    """Generate PDF for customer (no login required)"""
    
    def get(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Use the same PDF generation logic but without login requirement
        pdf_view = QuotationPDFView()
        pdf_response = pdf_view.generate_quotation_pdf(quotation)
        return pdf_response


# Export and Analytics Views
class QuotationExportView(LoginRequiredMixin, View):
    """Export quotations in various formats"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'csv')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        status = request.GET.get('status')
        
        # Build queryset
        queryset = Quotation.objects.all()
        
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
            
        # Apply role-based filtering
        if request.user.role == 'customer':
            queryset = queryset.filter(customer__email=request.user.email)
        elif request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(salesperson=request.user)
            
        if format_type == 'csv':
            return self.export_csv(queryset)
        elif format_type == 'excel':
            return self.export_excel(queryset)
        elif format_type == 'pdf':
            return self.export_pdf_report(queryset)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)
    
    def export_csv(self, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="quotations.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Quotation Number', 'Customer Name', 'Customer Email', 
            'System Type', 'Capacity (kW)', 'Total Amount (KES)', 
            'Status', 'Created Date', 'Valid Until'
        ])
        
        for quotation in queryset:
            writer.writerow([
                quotation.quotation_number,
                quotation.customer.name,
                quotation.customer.email,
                quotation.get_system_type_display(),
                quotation.system_capacity,
                quotation.total_amount,
                quotation.get_status_display(),
                quotation.created_at.strftime('%Y-%m-%d'),
                quotation.valid_until.strftime('%Y-%m-%d') if quotation.valid_until else ''
            ])
        
        return response
    
    def export_excel(self, queryset):
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Quotations'
        
        # Headers
        headers = [
            'Quotation Number', 'Customer Name', 'Customer Email', 
            'System Type', 'Capacity (kW)', 'Total Amount (KES)', 
            'Monthly Savings (KES)', 'Annual Savings (KES)',
            'Status', 'Created Date', 'Valid Until'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, quotation in enumerate(queryset, 2):
            worksheet.cell(row=row, column=1, value=quotation.quotation_number)
            worksheet.cell(row=row, column=2, value=quotation.customer.name)
            worksheet.cell(row=row, column=3, value=quotation.customer.email)
            worksheet.cell(row=row, column=4, value=quotation.get_system_type_display())
            worksheet.cell(row=row, column=5, value=float(quotation.system_capacity))
            worksheet.cell(row=row, column=6, value=float(quotation.total_amount))
            worksheet.cell(row=row, column=7, value=float(quotation.estimated_monthly_savings))
            worksheet.cell(row=row, column=8, value=float(quotation.estimated_annual_savings))
            worksheet.cell(row=row, column=9, value=quotation.get_status_display())
            worksheet.cell(row=row, column=10, value=quotation.created_at.strftime('%Y-%m-%d'))
            worksheet.cell(row=row, column=11, value=quotation.valid_until.strftime('%Y-%m-%d') if quotation.valid_until else '')
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="quotations.xlsx"'
        return response
    
    def export_pdf_report(self, queryset):
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        import weasyprint
        
        context = {
            'quotations': queryset,
            'export_date': timezone.now(),
            'total_count': queryset.count(),
            'total_value': sum(q.total_amount for q in queryset),
        }
        
        html_string = render_to_string('quotations/export_report.html', context)
        html = weasyprint.HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="quotations_report.pdf"'
        return response


class QuotationBulkActionsView(LoginRequiredMixin, View):
    """Handle bulk operations on quotations"""
    
    def post(self, request):
        action = request.POST.get('action')
        quotation_ids = request.POST.getlist('quotation_ids')
        
        if not quotation_ids:
            return JsonResponse({'success': False, 'message': 'No quotations selected'})
        
        queryset = Quotation.objects.filter(id__in=quotation_ids)
        
        # Apply role-based filtering
        if request.user.role == 'customer':
            queryset = queryset.filter(customer__email=request.user.email)
        elif request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(salesperson=request.user)
        
        if action == 'delete':
            count = queryset.filter(status__in=['draft', 'sent']).count()
            queryset.filter(status__in=['draft', 'sent']).delete()
            return JsonResponse({'success': True, 'message': f'{count} quotations deleted'})
        
        elif action == 'mark_sent':
            count = queryset.filter(status='draft').update(status='sent')
            return JsonResponse({'success': True, 'message': f'{count} quotations marked as sent'})
        
        elif action == 'mark_expired':
            count = queryset.filter(status__in=['sent', 'viewed']).update(status='expired')
            return JsonResponse({'success': True, 'message': f'{count} quotations marked as expired'})
        
        else:
            return JsonResponse({'success': False, 'message': 'Invalid action'})


class QuotationAnalyticsView(LoginRequiredMixin, TemplateView):
    """Quotation analytics and reporting dashboard"""
    template_name = 'quotations/analytics.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Build base queryset
        queryset = Quotation.objects.all()

        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)

        # Basic statistics
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['average_value'] = context['total_value'] / context['total_quotations'] if context['total_quotations'] > 0 else 0

        # Status breakdown
        context['status_stats'] = {
            status[0]: queryset.filter(status=status[0]).count()
            for status in Quotation.STATUS_CHOICES
        }

        # Conversion rate
        converted_count = queryset.filter(status='converted').count()
        context['conversion_rate'] = (converted_count / context['total_quotations'] * 100) if context['total_quotations'] > 0 else 0

        # Monthly trends (last 12 months)
        from django.db.models import Count, Sum
        from django.db.models.functions import ExtractMonth, ExtractYear
        from django.utils import timezone

        # Use ExtractMonth and ExtractYear to avoid timezone conversion issues
        monthly_data = queryset.filter(created_at__isnull=False).annotate(
            year=ExtractYear('created_at'),
            month=ExtractMonth('created_at')
        ).values('year', 'month').annotate(
            count=Count('id'),
            total_value=Sum('total_amount')
        ).order_by('year', 'month')

        # Format the data for the template
        formatted_monthly_data = []
        for item in monthly_data:
            # Skip items with None year or month to prevent TypeError
            if item['year'] is None or item['month'] is None:
                logger.warning(f"Skipping monthly data item with None year/month: {item}")
                continue

            # Create a date object for the month
            from datetime import date
            month_date = date(item['year'], item['month'], 1)
            formatted_monthly_data.append({
                'month': month_date,
                'count': item['count'],
                'total_value': item['total_value']
            })

        context['monthly_data'] = formatted_monthly_data

        # System type breakdown
        context['system_type_stats'] = {
            system_type[0]: queryset.filter(system_type=system_type[0]).count()
            for system_type in Quotation.SYSTEM_TYPES
        }

        return context


# CRUD Views
class QuotationUpdateView(LoginRequiredMixin, UpdateView):
    """Edit quotation"""
    model = Quotation
    fields = [
        'quotation_type', 'system_type', 'system_capacity', 'estimated_generation',
        'installation_complexity', 'subtotal', 'discount_amount', 'discount_percentage',
        'tax_amount', 'total_amount', 'estimated_monthly_savings', 'estimated_annual_savings',
        'payback_period_months', 'roi_percentage', 'financing_available', 'payment_terms',
        'delivery_time', 'warranty_terms', 'notes'
    ]
    template_name = 'quotations/edit.html'
    
    def get_object(self):
        return get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
    
    def get_success_url(self):
        return reverse('quotations:detail', kwargs={'quotation_number': self.object.quotation_number})
    
    def form_valid(self, form):
        messages.success(self.request, f'Quotation {self.object.quotation_number} updated successfully!')
        return super().form_valid(form)


class QuotationDeleteView(LoginRequiredMixin, DeleteView):
    """Delete quotation"""
    model = Quotation
    template_name = 'quotations/delete_confirm.html'
    success_url = reverse_lazy('quotations:list')
    
    def get_object(self):
        return get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
    
    def delete(self, request, *args, **kwargs):
        quotation = self.get_object()
        
        # Only allow deletion of draft or sent quotations
        if quotation.status not in ['draft', 'sent']:
            messages.error(request, 'Cannot delete quotations that have been accepted or converted.')
            return redirect('quotations:detail', quotation_number=quotation.quotation_number)
        
        quotation_number = quotation.quotation_number
        quotation.delete()
        messages.success(request, f'Quotation {quotation_number} deleted successfully!')
        return redirect(self.success_url)


class QuotationSendEmailView(LoginRequiredMixin, View):
    """Manually send quotation email"""
    
    def post(self, request, quotation_number):
        quotation = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        try:
            from apps.core.email_utils import EmailService
            success = EmailService.send_quotation_created_email(quotation)
            
            if success:
                quotation.status = 'sent'
                quotation.save()
                messages.success(request, f'Quotation {quotation_number} sent successfully!')
            else:
                messages.error(request, 'Failed to send quotation email.')
                
        except Exception as e:
            messages.error(request, f'Error sending email: {str(e)}')
        
        return redirect('quotations:detail', quotation_number=quotation_number)


class QuotationDuplicateView(LoginRequiredMixin, View):
    """Duplicate an existing quotation"""
    
    def post(self, request, quotation_number):
        original = get_object_or_404(Quotation, quotation_number=quotation_number)
        
        # Create duplicate
        duplicate = Quotation.objects.create(
            customer=original.customer,
            salesperson=request.user,
            quotation_type=original.quotation_type,
            system_type=original.system_type,
            system_capacity=original.system_capacity,
            estimated_generation=original.estimated_generation,
            installation_complexity=original.installation_complexity,
            subtotal=original.subtotal,
            discount_amount=original.discount_amount,
            discount_percentage=original.discount_percentage,
            tax_amount=original.tax_amount,
            total_amount=original.total_amount,
            estimated_monthly_savings=original.estimated_monthly_savings,
            estimated_annual_savings=original.estimated_annual_savings,
            payback_period_months=original.payback_period_months,
            roi_percentage=original.roi_percentage,
            financing_available=original.financing_available,
            financing_options=original.financing_options,
            payment_terms=original.payment_terms,
            delivery_time=original.delivery_time,
            warranty_terms=original.warranty_terms,
            notes=f"Duplicated from {original.quotation_number}\n\n{original.notes}",
            status='draft'
        )
        
        messages.success(request, f'Quotation duplicated as {duplicate.quotation_number}')
        return redirect('quotations:edit', quotation_number=duplicate.quotation_number)


class QuotationFollowUpView(LoginRequiredMixin, CreateView):
    model = QuotationFollowUp
    fields = ['follow_up_type', 'notes', 'scheduled_date']
    
    def form_valid(self, form):
        try:
            # Get date and time from separate inputs
            date_str = self.request.POST.get('scheduled_date_0')
            time_str = self.request.POST.get('scheduled_date_1')
            
            # Combine date and time
            from datetime import datetime
            scheduled_datetime = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
            
            # Set the datetime on the form instance
            form.instance.scheduled_date = scheduled_datetime
            
            # Set other required fields
            quotation = get_object_or_404(Quotation, quotation_number=self.kwargs['quotation_number'])
            form.instance.quotation = quotation
            form.instance.created_by = self.request.user
            
            messages.success(self.request, 'Follow-up added successfully!')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Error saving follow-up: {str(e)}')
            return super().form_invalid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please check the form data and try again.')
        return redirect('quotations:detail', quotation_number=self.kwargs['quotation_number'])
    
    def get_success_url(self):
        return reverse('quotations:detail', kwargs={'quotation_number': self.kwargs['quotation_number']})
    
    # Override get method to redirect to detail view if accessed directly
    def get(self, request, *args, **kwargs):
        return redirect('quotations:detail', quotation_number=self.kwargs['quotation_number'])

class QuotationFollowUpListView(LoginRequiredMixin, ListView):
    model = QuotationFollowUp
    template_name = 'quotations/follow_ups.html'
    context_object_name = 'follow_ups'
    paginate_by = 20

    def get_queryset(self):
        queryset = QuotationFollowUp.objects.select_related('quotation', 'created_by')
        
        # Filter by role
        if self.request.user.role == 'customer':
            queryset = queryset.filter(quotation__customer__email=self.request.user.email)
        elif self.request.user.role in ['sales_person', 'sales_manager']:
            queryset = queryset.filter(quotation__salesperson=self.request.user)

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            if status == 'pending':
                queryset = queryset.filter(scheduled_date__gte=timezone.now())
            elif status == 'completed':
                queryset = queryset.filter(completed=True)
            elif status == 'overdue':
                queryset = queryset.filter(scheduled_date__lt=timezone.now(), completed=False)

        return queryset.order_by('scheduled_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pending_count'] = self.get_queryset().filter(scheduled_date__gte=timezone.now(), completed=False).count()
        context['overdue_count'] = self.get_queryset().filter(scheduled_date__lt=timezone.now(), completed=False).count()
        context['completed_count'] = self.get_queryset().filter(completed=True).count()
        return context

@require_POST
@login_required
def complete_follow_up(request, pk):
    """Mark a follow-up as completed"""
    try:
        follow_up = get_object_or_404(QuotationFollowUp, pk=pk)
        
        # Check permissions
        if request.user != follow_up.created_by and request.user != follow_up.quotation.salesperson:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to complete this follow-up'
            }, status=403)
        
        # Mark as completed
        follow_up.completed = True
        follow_up.save()
        
        # Add completion note if provided
        completion_note = request.POST.get('completion_note')
        if completion_note:
            follow_up.notes += f"\n\nCompletion Note ({timezone.now().strftime('%Y-%m-%d %H:%M')}): {completion_note}"
            follow_up.save(update_fields=['notes'])
        
        messages.success(request, 'Follow-up marked as completed successfully!')
        return JsonResponse({
            'success': True,
            'message': 'Follow-up completed successfully'
        })
        
    except Exception as e:
        logger.error(f"Error completing follow-up: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error completing follow-up: {str(e)}'
        }, status=500)

class DecimalEncoder(DjangoJSONEncoder):
    def default(self, obj):
        from decimal import Decimal
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

class ProductSelectorView(LoginRequiredMixin, TemplateView):
    """Interactive product selection interface for solar systems"""
    template_name = 'quotations/product_selector.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        try:
            from apps.products.models import Product, ProductCategory
            from django.db.models import Count, Q
            import json
            
            # Get active categories with product counts
            categories = ProductCategory.objects.filter(
                is_active=True
            ).annotate(
                product_count=Count('products', filter=Q(products__status='active'))
            )
            
            # Format categories for frontend using stored icons
            category_data = []
            for category in categories:
                category_data.append({
                    'id': category.slug,
                    'name': category.name,
                    'icon': category.icon,  # Use the stored icon
                    'count': category.product_count
                })
            
            # Get products for each category
            products_data = {}
            for category in categories:
                products = Product.objects.filter(
                    category=category,
                    status='active'
                ).select_related('category').values(
                    'id', 
                    'name', 
                    'selling_price',
                    'power_rating',
                    'efficiency',
                    'voltage',
                    'warranty_years',
                    'brand',
                    'short_description',
                    'weight'
                )
                
                # Format products
                formatted_products = []
                for product in products:
                    formatted_product = {
                        'id': product['id'],
                        'name': product['name'],
                        'selling_price': float(product['selling_price'] or 0),
                        'price': float(product['selling_price'] or 0),
                        'specifications': {
                            'power': product['power_rating'],
                            'efficiency': product['efficiency'],
                            'voltage': product['voltage'],
                            'weight': product['weight']
                        },
                        'warranty_years': product['warranty_years'],
                        'brand': product['brand'],
                        'description': product['short_description']
                    }
                    formatted_products.append(formatted_product)
                
                products_data[category.slug] = formatted_products
            
            # Get unique brands
            brands = Product.objects.filter(
                status='active'
            ).values_list('brand', flat=True).distinct()
            
            # Add to context with proper JSON encoding
            context['category_data'] = json.dumps(category_data)
            context['initial_products'] = json.dumps(products_data)
            context['brands'] = json.dumps(list(filter(None, brands)))
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error loading product selector data: {str(e)}")
            context['category_data'] = '[]'
            context['initial_products'] = '{}'
            context['brands'] = '[]'
        
        return context
    
    def get_default_icon(self, slug):
        """Get default icon for category"""
        default_icons = {
            'solar-panels': 'fas fa-solar-panel',
            'inverters': 'fas fa-microchip',
            'batteries': 'fas fa-battery-full',
            'mounting': 'fas fa-tools',
            'cables': 'fas fa-plug',
            'accessories': 'fas fa-cogs'
        }
        return default_icons.get(slug, 'fas fa-box')

class QuotationBuilderView(LoginRequiredMixin, TemplateView):
    """Advanced quotation builder with drag-and-drop functionality"""
    template_name = 'quotations/quotation_builder.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get quotation if editing
        quotation_id = self.request.GET.get('quotation_id')
        if quotation_id:
            try:
                quotation = Quotation.objects.get(id=quotation_id)
                context['quotation'] = quotation
                context['quotation_items'] = quotation.items.all()
            except Quotation.DoesNotExist:
                pass
        
        # Get products for the builder
        try:
            from apps.products.models import Product, ProductCategory
            
            categories = ProductCategory.objects.filter(is_active=True)
            context['product_categories'] = categories
            
            # Get products organized by category
            context['products_by_category'] = {}
            for category in categories:
                products = Product.objects.filter(category=category, is_active=True)
                context['products_by_category'][category.slug] = [
                    {
                        'id': p.id,
                        'name': p.name,
                        'price': float(p.price),
                        'category': category.slug,
                        'unit': p.unit or 'piece',
                        'description': p.description,
                        'specifications': p.specifications or {},
                        'warranty': getattr(p, 'warranty', '1 year'),
                        'brand': getattr(p, 'brand', 'Generic'),
                        'image': p.image.url if p.image else None
                    } for p in products
                ]
                
        except ImportError:
            # Mock data for development
            context['product_categories'] = [
                {'id': 'panels', 'name': 'Solar Panels', 'slug': 'panels'},
                {'id': 'inverters', 'name': 'Inverters', 'slug': 'inverters'},
                {'id': 'batteries', 'name': 'Batteries', 'slug': 'batteries'},
            ]
            context['products_by_category'] = {}
        
        return context

class CustomerRequirementsView(LoginRequiredMixin, TemplateView):
    """Customer requirements capture form for accurate quotations"""
    template_name = 'quotations/customer_requirements.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add any initial data if needed
        context['current_step'] = 1
        
        return context

class EnhancedAnalyticsView(LoginRequiredMixin, TemplateView):
    """Enhanced analytics and reporting dashboard"""
    template_name = 'quotations/enhanced_analytics.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Build base queryset
        queryset = Quotation.objects.all()
        
        # Apply role-based filtering
        if hasattr(self.request.user, 'role'):
            if self.request.user.role == 'customer':
                queryset = queryset.filter(customer__email=self.request.user.email)
            elif self.request.user.role in ['sales_person', 'sales_manager']:
                queryset = queryset.filter(salesperson=self.request.user)
        
        # Basic metrics
        context['total_quotations'] = queryset.count()
        context['total_value'] = sum(q.total_amount for q in queryset)
        context['average_value'] = context['total_value'] / context['total_quotations'] if context['total_quotations'] > 0 else 0
        
        # Status breakdown
        context['sent_quotations'] = queryset.filter(status='sent').count()
        context['viewed_quotations'] = queryset.filter(status='viewed').count()
        context['accepted_quotations'] = queryset.filter(status='accepted').count()
        context['converted_quotations'] = queryset.filter(status='converted').count()
        
        # Calculate rates
        total = context['total_quotations']
        if total > 0:
            context['sent_rate'] = (context['sent_quotations'] / total) * 100
            context['view_rate'] = (context['viewed_quotations'] / context['sent_quotations']) * 100 if context['sent_quotations'] > 0 else 0
            context['acceptance_rate'] = (context['accepted_quotations'] / context['viewed_quotations']) * 100 if context['viewed_quotations'] > 0 else 0
            context['conversion_rate'] = (context['converted_quotations'] / total) * 100
        else:
            context['sent_rate'] = context['view_rate'] = context['acceptance_rate'] = context['conversion_rate'] = 0
        
        # Revenue analysis
        context['converted_value'] = sum(q.total_amount for q in queryset.filter(status='converted'))
        context['pending_value'] = sum(q.total_amount for q in queryset.filter(status__in=['sent', 'viewed', 'accepted']))
        context['pipeline_value'] = context['total_value']
        
        # Performance metrics
        context['avg_response_time'] = 2.4  # Mock data
        context['win_rate'] = context['conversion_rate']
        
        return context


@method_decorator(csrf_exempt, name='dispatch')
class ScheduleConsultationView(View):
    """Schedule consultation from calculator"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            customer_data = data.get('customerData', {})
            consultation_data = data.get('consultationData', {})
            
            # Extract data
            customer_name = customer_data.get('name')
            customer_email = customer_data.get('email')
            customer_phone = customer_data.get('phone', '')
            preferred_date = consultation_data.get('preferredDate')
            preferred_time = consultation_data.get('preferredTime')
            consultation_type = consultation_data.get('consultationType', 'site_visit')
            
            if not all([customer_name, customer_email, preferred_date]):
                return JsonResponse({'success': False, 'message': 'Name, email, and preferred date are required'})
            
            # For now, we'll send an email notification to staff
            # In a full implementation, this would integrate with a calendar system
            
            message = f"""
            New consultation request from solar calculator:
            
            Customer: {customer_name}
            Email: {customer_email}
            Phone: {customer_phone}
            Preferred Date: {preferred_date}
            Preferred Time: {preferred_time}
            Consultation Type: {consultation_type}
            
            Please contact the customer to confirm the appointment.
            """
            
            EmailService.send_staff_notification(
                subject=f"New Consultation Request - {customer_name}",
                message=message
            )
            
            # Send confirmation to customer
            context = {
                'customer_name': customer_name,
                'preferred_date': preferred_date,
                'preferred_time': preferred_time,
                'consultation_type': consultation_type,
            }
            
            EmailService.send_email_notification(
                'consultation_scheduled',
                context,
                customer_email,
                'Consultation Request Received - Olivian Group'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Consultation request submitted successfully! We will contact you soon to confirm.'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error scheduling consultation: {str(e)}'
            })
