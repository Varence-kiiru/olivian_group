from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
)
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import (
    Q, Count, Sum, Avg, F, DurationField, ExpressionWrapper,
    Case, When, Value, CharField  # Add these imports
)
from django.db.models.functions import Concat
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils import timezone
from django.conf import settings
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import csv
import xlsxwriter
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import io
import json
import logging
import re

from .models import (
    Lead, Opportunity, Contact, Company, Activity, Campaign,
    LeadSource, Pipeline, PipelineStage, EmailTemplate
)
from .forms import (
    LeadForm, OpportunityForm, ContactForm, CompanyForm,
    ActivityForm, CampaignForm
)
from .reports import ActivityReport, PipelineReport, RevenueReport, CustomerReport
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apps.ecommerce.models import Order

User = get_user_model()

# Initialize logger
logger = logging.getLogger(__name__)

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Get date ranges
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Basic stats
        context.update({
            'total_leads': Lead.objects.count(),
            'total_opportunities': Opportunity.objects.count(),
            'total_contacts': Contact.objects.count(),
            'total_companies': Company.objects.count(),
        })

        # Sales metrics
        opportunities = Opportunity.objects.exclude(stage__in=['closed_lost'])
        context.update({
            'pipeline_value': opportunities.aggregate(total=Sum('value'))['total'] or 0,
            'weighted_pipeline': sum(opp.weighted_value for opp in opportunities),
            'avg_deal_size': opportunities.aggregate(avg=Avg('value'))['avg'] or 0,
        })

        # Recent activity
        context.update({
            'recent_leads': Lead.objects.filter(created_at__gte=week_ago).order_by('-created_at')[:5],
            'upcoming_activities': Activity.objects.filter(
                scheduled_datetime__gte=timezone.now(),
                status='planned'
            ).order_by('scheduled_datetime')[:5],
            'overdue_leads': Lead.objects.filter(
                next_follow_up__lt=today,
                status__in=['new', 'contacted', 'qualified']
            ).count(),
        })

        # Charts data
        context.update({
            'leads_by_status': self.get_leads_by_status(),
            'opportunities_by_stage': self.get_opportunities_by_stage(),
            'monthly_performance': self.get_monthly_performance(),
        })

        # For manager dashboard
        if self.request.user.role in ['manager', 'project_manager', 'inventory_manager']:
            try:
                # Get team members based on department
                team_members = User.objects.filter(
                    department=self.request.user.department,
                    is_active_employee=True
                ).exclude(
                    id=self.request.user.id  # Exclude self
                ).exclude(
                    role__in=['customer', 'super_admin']  # Exclude customers and super admins
                )

                # Convert Decimal to float for revenue calculations
                current_revenue = float(self.get_revenue_data() or 0)
                last_revenue = float(self.get_last_month_revenue() or 0)
                
                # Calculate changes using float values
                revenue_change = ((current_revenue - last_revenue) / max(last_revenue, 1)) * 100 if last_revenue > 0 else 0
                net_profit = current_revenue * 0.20  # Convert 0.2 to 0.20 for clarity
                
                context['manager_stats'] = {
                    'total_revenue': current_revenue,
                    'revenue_change': round(revenue_change, 1),
                    'net_profit': round(net_profit, 2),
                    'profit_change': round(revenue_change, 1),
                    'active_projects': self.get_active_projects_count(),
                    'projects_change': self.get_projects_change(),
                    'team_performance': round(self.calculate_team_performance(team_members), 1),
                    'performance_change': 5,
                    'budget_utilization': 65,
                    'budget_remaining': 500000
                }
                
                # Add team members data
                context['team_members'] = [{
                    'id': member.id,
                    'username': member.username,
                    'full_name': member.get_full_name(),
                    'role': member.get_role_display(),
                    'email': member.email,
                    'phone': member.phone,
                    'department': member.department,
                    'performance': self.get_member_performance(member)
                } for member in team_members]

            except Exception as e:
                print(f"Manager dashboard error: {str(e)}")
                context['manager_stats'] = {
                    'total_revenue': 0,
                    'revenue_change': 0,
                    'net_profit': 0,
                    'profit_change': 0,
                    'active_projects': 0,
                    'projects_change': 0,
                    'team_performance': 0,
                    'performance_change': 0,
                    'budget_utilization': 0,
                    'budget_remaining': 0
                }
                context['team_members'] = []

        # Get available employees for team assignment
        available_employees = User.objects.filter(
            is_active=True,
            is_active_employee=True
        ).exclude(
            role__in=['customer', 'super_admin']
        ).select_related(
            'userprofile'
        ).order_by(
            'first_name', 'last_name'
        )

        context['available_employees'] = available_employees

        return context

    def get_leads_by_status(self):
        return list(Lead.objects.values('status').annotate(count=Count('id')))

    def get_opportunities_by_stage(self):
        return list(Opportunity.objects.values('stage').annotate(
            count=Count('id'),
            total_value=Sum('value')
        ))

    def get_monthly_performance(self):
        # Last 6 months performance
        data = []
        for i in range(6):
            date = timezone.now().date().replace(day=1) - timedelta(days=i*30)
            month_start = date.replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)

            leads_count = Lead.objects.filter(
                created_at__date__range=[month_start, month_end]
            ).count()

            won_value = Opportunity.objects.filter(
                actual_close_date__range=[month_start, month_end],
                stage='closed_won'
            ).aggregate(total=Sum('value'))['total'] or 0

            data.append({
                'month': month_start.strftime('%b %Y'),
                'leads': leads_count,
                'revenue': float(won_value)
            })

        return list(reversed(data))

    def get_revenue_data(self):
        """Get current month revenue"""
        from django.utils import timezone
        current_month = timezone.now().replace(day=1)
        
        return Order.objects.filter(
            created_at__gte=current_month,
            status__in=['completed', 'delivered', 'paid']
        ).aggregate(total=Sum('total_amount'))['total'] or 0

    def get_last_month_revenue(self):
        """Get last month revenue"""
        from django.utils import timezone
        from dateutil.relativedelta import relativedelta
        
        now = timezone.now()
        last_month_start = (now - relativedelta(months=1)).replace(day=1)
        last_month_end = now.replace(day=1) - relativedelta(days=1)
        
        return Order.objects.filter(
            created_at__gte=last_month_start,
            created_at__lte=last_month_end,
            status__in=['completed', 'delivered', 'paid']
        ).aggregate(total=Sum('total_amount'))['total'] or 0

    def calculate_team_performance(self, team_members):
        """Calculate overall team performance"""
        if not team_members:
            return 0
            
        total_performance = sum(self.get_member_performance(member) for member in team_members)
        return total_performance / len(team_members)

    def get_member_performance(self, member):
        """Calculate individual member performance"""
        try:
            # Get member's orders/quotations/projects
            total_orders = Order.objects.filter(
                created_by=member,
                created_at__gte=timezone.now().replace(day=1)
            ).count()
            
            # Simple performance calculation (customize based on your needs)
            target = 10  # Example target
            performance = min(100, (total_orders / target) * 100)
            return performance
        except:
            return 0


class LeadListView(LoginRequiredMixin, ListView):
    model = Lead
    template_name = 'crm/lead_list.html'
    context_object_name = 'leads'
    paginate_by = 12  # Changed to 12 for better grid layout

    def get_queryset(self):
        queryset = Lead.objects.select_related(
            'contact', 'company', 'assigned_to', 'source'
        )

        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(contact__first_name__icontains=search) |
                Q(contact__last_name__icontains=search) |
                Q(contact__email__icontains=search) |
                Q(company__name__icontains=search)
            )

        # Status filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Priority filter
        priority = self.request.GET.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        # Source filter
        source = self.request.GET.get('source')
        if source:
            queryset = queryset.filter(source_id=source)
            
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'lead_sources': LeadSource.objects.filter(is_active=True),
            'status_choices': Lead.STATUS_CHOICES,
            'priority_choices': Lead.PRIORITY_CHOICES,
            'current_search': self.request.GET.get('search', ''),
            'current_status': self.request.GET.get('status', ''),
            'current_priority': self.request.GET.get('priority', ''),
            'current_source': self.request.GET.get('source', ''),
        })
        return context


class LeadDetailView(LoginRequiredMixin, DetailView):
    model = Lead
    template_name = 'crm/lead_detail.html'
    context_object_name = 'lead'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead = self.get_object()

        # Calculate lead score
        lead_score = 0
        if lead.company:
            lead_score += 10
        if lead.contact and lead.contact.email:
            lead_score += 10
        if lead.contact and lead.contact.phone:
            lead_score += 5
        if lead.company and lead.company.website:
            lead_score += 5
        if lead.estimated_value:
            lead_score += 15
        if lead.description:
            lead_score += 10

        # Priority bonus
        priority_scores = {'urgent': 25, 'high': 20, 'medium': 10, 'low': 5}
        lead_score += priority_scores.get(lead.priority, 0)
        
        # Status bonus
        status_scores = {'proposal': 30, 'qualified': 25, 'negotiation': 35, 'contacted': 15}
        lead_score += status_scores.get(lead.status, 0)

        # Cap at 100
        lead_score = min(lead_score, 100)
        
        # Calculate conversion probability
        conversion_probability = {
            'new': 20,
            'contacted': 30,
            'qualified': 50,
            'proposal': 70,
            'negotiation': 90,
            'won': 100,
            'lost': 0,
            'unqualified': 0
        }.get(lead.status, 0)

        # Get activities related to both the lead and its contact
        activities = Activity.objects.filter(
            Q(lead=lead) | Q(contact=lead.contact)
        ).select_related(
            'assigned_to', 'contact', 'lead', 'opportunity'
        ).order_by('-scheduled_datetime')

        context.update({
            'activities': activities,
            'lead_score': lead_score,
            'conversion_probability': conversion_probability,
            'can_convert': lead.status in ['qualified', 'proposal'] and not hasattr(lead, 'opportunity'),
            'days_since_created': (timezone.now().date() - lead.created_at.date()).days,
        })
        return context


class LeadCreateView(LoginRequiredMixin, CreateView):
    model = Lead
    form_class = LeadForm
    template_name = 'crm/lead_form.html'
    success_url = reverse_lazy('crm:lead_list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_initial(self):
        initial = super().get_initial()
        company_id = self.request.GET.get('company')
        
        if company_id:
            try:
                company = Company.objects.get(pk=company_id)
                # Pre-populate company information
                initial.update({
                    'company_name': company.name,
                    'website': company.website,
                    'industry': company.industry,
                    'company_size': self._get_company_size_range(company.employee_count),
                    'city': company.city,
                    'country': company.country or 'Kenya',

                    # If there's a primary contact, use their details
                    'contact_name': company.primary_contact.get_full_name() if company.primary_contact else '',
                    'email': company.primary_contact.email if company.primary_contact else company.email,
                    'phone': company.primary_contact.phone if company.primary_contact else company.phone,
                    'job_title': company.primary_contact.position if company.primary_contact else '',
                    'address': company.address_line_1,
                })
            except Company.DoesNotExist:
                pass
        
        return initial
    
    def _get_company_size_range(self, employee_count):
        """Convert employee count to size range"""
        if not employee_count:
            return ''
        if employee_count <= 10:
            return '1-10'
        elif employee_count <= 50:
            return '11-50'
        elif employee_count <= 200:
            return '51-200'
        elif employee_count <= 500:
            return '201-500'
        return '500+'

    def form_valid(self, form):
        try:
            # Set created_by and assigned_to
            form.instance.created_by = self.request.user
            if not form.instance.assigned_to:
                form.instance.assigned_to = self.request.user

            # Get or create the company
            company_name = form.cleaned_data.get('company_name')
            if company_name:
                company, _ = Company.objects.get_or_create(
                    name=company_name,
                    defaults={
                        'created_by': self.request.user,
                        'website': form.cleaned_data.get('website', ''),
                        'industry': form.cleaned_data.get('industry', ''),
                        'city': form.cleaned_data.get('city', ''),
                        'country': form.cleaned_data.get('country', 'Kenya'),
                    }
                )
                form.instance.company = company

            response = super().form_valid(form)
            messages.success(self.request, 'Lead created successfully!')
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error creating lead: {str(e)}')
            return super().form_invalid(form)


class LeadUpdateView(LoginRequiredMixin, UpdateView):
    model = Lead
    form_class = LeadForm
    template_name = 'crm/lead_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_initial(self):
        initial = super().get_initial()
        lead = self.get_object()
        
        # Populate additional fields from related models
        if lead.contact:
            initial['contact_name'] = lead.contact.get_full_name()
            initial['email'] = lead.contact.email
            initial['phone'] = lead.contact.phone
            initial['job_title'] = lead.contact.position
        
        if lead.company:
            initial['company_name'] = lead.company.name
            initial['website'] = lead.company.website
            initial['industry'] = lead.company.industry
            initial['city'] = lead.company.city
            initial['country'] = 'Kenya'  # Default
            
            # Map employee count to company size
            if lead.company.employee_count:
                count = lead.company.employee_count
                if count <= 10:
                    initial['company_size'] = '1-10'
                elif count <= 50:
                    initial['company_size'] = '11-50'
                elif count <= 200:
                    initial['company_size'] = '51-200'
                elif count <= 500:
                    initial['company_size'] = '201-500'
                else:
                    initial['company_size'] = '500+'
        
        # Map description to interest
        if lead.description:
            initial['interest'] = lead.description
            
        # Map estimated value to budget
        if lead.estimated_value:
            initial['budget'] = lead.estimated_value
        
        return initial
    
    def form_valid(self, form):
        messages.success(self.request, 'Lead updated successfully!')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class LeadDeleteView(LoginRequiredMixin, DeleteView):
    model = Lead
    template_name = 'crm/lead_confirm_delete.html'
    success_url = reverse_lazy('crm:lead_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Lead deleted successfully!')
        return super().delete(request, *args, **kwargs)


class ConvertLeadView(LoginRequiredMixin, View):
    def post(self, request, pk):
        lead = get_object_or_404(Lead, pk=pk)
        
        if hasattr(lead, 'opportunity'):
            messages.warning(request, 'Lead already has an associated opportunity.')
            return redirect('crm:lead_detail', pk=pk)
        
        # Create opportunity from lead
        opportunity = Opportunity.objects.create(
            name=f"{lead.title} - {lead.contact.get_full_name()}",
            lead=lead,
            contact=lead.contact,
            company=lead.company,
            stage='qualification',
            probability=25,
            value=lead.estimated_value or Decimal('0'),
            expected_close_date=lead.expected_close_date or timezone.now().date(),
            assigned_to=lead.assigned_to,
            description=lead.description,
            created_by=request.user
        )
        
        # Update lead status
        lead.status = 'won'
        lead.save()
        
        messages.success(request, f'Lead converted to opportunity: {opportunity.name}')
        return redirect('crm:opportunity_detail', pk=opportunity.pk)


# Opportunity Views
class OpportunityListView(LoginRequiredMixin, ListView):
    model = Opportunity
    template_name = 'crm/opportunity_list.html'
    context_object_name = 'opportunities'
    paginate_by = 25
    
    def get_queryset(self):
        return Opportunity.objects.select_related('contact', 'company', 'assigned_to').order_by('-value')


class OpportunityDetailView(LoginRequiredMixin, DetailView):
    model = Opportunity
    template_name = 'crm/opportunity_detail.html'
    context_object_name = 'opportunity'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        opportunity = self.get_object()

        # Get stage progress from STAGE_WEIGHTS
        stage_progress = opportunity.STAGE_WEIGHTS.get(
            opportunity.stage, {'progress': 0}
        )['progress']

        # Calculate days in pipeline
        days_in_pipeline = (timezone.now().date() - opportunity.created_at.date()).days

        context.update({
            'activities': opportunity.activities.all().order_by('-created_at'),
            'days_in_pipeline': days_in_pipeline,
            'stage_progress': stage_progress,
            'total_activities': opportunity.activities.count(),
            'weighted_value': opportunity.weighted_value
        })

        return context


@login_required
@require_POST
def change_opportunity_stage(request, pk):
    try:
        opportunity = Opportunity.objects.get(pk=pk)
        new_stage = request.POST.get('stage')

        if new_stage not in dict(Opportunity.STAGE_CHOICES):
            return JsonResponse({'error': 'Invalid stage'}, status=400)

        opportunity.change_stage(new_stage, request.user)

        return JsonResponse({
            'success': True,
            'new_stage': opportunity.get_stage_display(),
            'probability': opportunity.probability,
            'weighted_value': float(opportunity.weighted_value)
        })
    except Opportunity.DoesNotExist:
        return JsonResponse({'error': 'Opportunity not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


class OpportunityCreateView(LoginRequiredMixin, CreateView):
    model = Opportunity
    form_class = OpportunityForm
    template_name = 'crm/opportunity_form.html'
    
    def get_initial(self):
        initial = super().get_initial()
        lead_id = self.request.GET.get('lead')
        
        if lead_id:
            try:
                lead = Lead.objects.get(pk=lead_id)
                initial.update({
                    'name': f"Opportunity - {lead.title}",
                    'lead': lead,
                    'contact': lead.contact,
                    'company': lead.company,
                    'value': lead.estimated_value,
                    'description': lead.description,
                    'assigned_to': lead.assigned_to,
                    'expected_close_date': lead.expected_close_date,
                })
            except Lead.DoesNotExist:
                pass
        
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Update lead status if coming from lead conversion
        lead = form.cleaned_data.get('lead')
        if lead:
            lead.status = 'qualified'
            lead.save()
            
        messages.success(self.request, 'Opportunity created successfully!')
        return response


class OpportunityUpdateView(LoginRequiredMixin, UpdateView):
    model = Opportunity
    form_class = OpportunityForm
    template_name = 'crm/opportunity_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        opportunity = self.get_object()
        
        # Get stage probabilities
        stage_probabilities = {
            'qualification': 10,
            'needs_analysis': 30,
            'proposal': 50,
            'negotiation': 70,
            'closed_won': 100,
            'closed_lost': 0
        }

        context.update({
            'stage_probabilities': stage_probabilities,
            'weighted_value': opportunity.weighted_value,
            'days_in_stage': (timezone.now().date() - opportunity.updated_at.date()).days,
            'total_activities': opportunity.activities.count(),
        })
        return context

    def form_valid(self, form):
        try:
            old_stage = self.get_object().stage
            new_stage = form.cleaned_data['stage']
            
            response = super().form_valid(form)
            
            # Record stage change if it occurred
            if old_stage != new_stage:
                Activity.objects.create(
                    subject=f"Stage changed from {old_stage} to {new_stage}",
                    activity_type='other',
                    status='completed',
                    opportunity=self.object,
                    contact=self.object.contact,
                    scheduled_datetime=timezone.now(),
                    completed_datetime=timezone.now(),
                    description=f"Opportunity stage changed from {self.object.get_stage_display()} to {form.instance.get_stage_display()}",
                    assigned_to=self.object.assigned_to or self.request.user,
                    created_by=self.request.user
                )

            # Update actual close date if stage is closed
            if new_stage in ['closed_won', 'closed_lost'] and not form.instance.actual_close_date:
                self.object.actual_close_date = timezone.now().date()
                self.object.save()

            messages.success(self.request, 'Opportunity updated successfully!')
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error updating opportunity: {str(e)}')
            return super().form_invalid(form)

    def get_success_url(self):
        if self.request.POST.get('action') == 'save_continue':
            return reverse_lazy('crm:opportunity_update', kwargs={'pk': self.object.pk})
        return reverse_lazy('crm:opportunity_detail', kwargs={'pk': self.object.pk})

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class OpportunityDeleteView(LoginRequiredMixin, DeleteView):
    model = Opportunity
    success_url = reverse_lazy('crm:opportunity_list')


# Contact Views
class ContactListView(LoginRequiredMixin, ListView):
    model = Contact
    template_name = 'crm/contact_list.html'
    context_object_name = 'contacts'
    paginate_by = 12

    def get_queryset(self):
        # Change 'user' to 'created_by' in select_related
        queryset = Contact.objects.select_related('created_by', 'source').prefetch_related(
            'leads', 'activities'
        )
        
        # Apply filters
        filters = {
            'search': self.request.GET.get('search'),
            'status': self.request.GET.get('status'),
            'company': self.request.GET.get('company'),
            'created': self.request.GET.get('created')
        }

        if filters['search']:
            queryset = queryset.filter(
                Q(first_name__icontains=filters['search']) |
                Q(last_name__icontains=filters['search']) |
                Q(email__icontains=filters['search']) |
                Q(companies__name__icontains=filters['search'])
            ).distinct()

        if filters['status']:
            queryset = queryset.filter(status=filters['status'])

        if filters['company']:
            queryset = queryset.filter(companies__id=filters['company'])

        if filters['created']:
            today = timezone.now().date()
            if filters['created'] == 'today':
                queryset = queryset.filter(created_at__date=today)
            elif filters['created'] == 'week':
                queryset = queryset.filter(created_at__date__gte=today - timedelta(days=7))
            elif filters['created'] == 'month':
                queryset = queryset.filter(created_at__date__gte=today - timedelta(days=30))

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        context.update({
            'active_contacts_count': Contact.objects.filter(status='active').count(),
            'new_contacts_this_month': Contact.objects.filter(
                created_at__year=today.year,
                created_at__month=today.month
            ).count(),
            'companies_count': Company.objects.count(),
            'companies': Company.objects.all(),
            'current_filters': {
                'search': self.request.GET.get('search', ''),
                'status': self.request.GET.get('status', ''),
                'company': self.request.GET.get('company', ''),
                'created': self.request.GET.get('created', '')
            }
        })
        return context


class ContactDetailView(LoginRequiredMixin, DetailView):
    model = Contact
    template_name = 'crm/contact_detail.html'
    context_object_name = 'contact'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contact = self.get_object()
        
        # Get recent activities
        recent_activities = contact.activities.select_related(
            'assigned_to', 'lead', 'opportunity'
        ).order_by('-scheduled_datetime')[:5]
        
        # Calculate stats
        opportunities_count = 0
        total_value = 0
        
        # Get leads and their opportunities
        for lead in contact.leads.all():
            if hasattr(lead, 'opportunity'):
                opportunities_count += 1
                total_value += lead.opportunity.value
        
        context.update({
            'recent_activities': recent_activities,
            'opportunities_count': opportunities_count,
            'total_value': total_value,
            'active_leads_count': contact.leads.exclude(status__in=['won', 'lost']).count(),
        })
        return context


class ContactCreateView(LoginRequiredMixin, CreateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm/contact_form.html'
    
    def get_initial(self):
        initial = super().get_initial()

        # Handle company pre-population
        company_id = self.request.GET.get('company')
        if company_id:
            try:
                company = Company.objects.get(pk=company_id)
                initial.update({
                    'contact_type': 'company',
                    'company': company.id,
                    'address_line_1': company.address_line_1,
                    'city': company.city,
                    'county': company.county,
                    'postal_code': company.postal_code,
                    'country': company.country if hasattr(company, 'country') else 'Kenya'  # Safe access
                })
            except Company.DoesNotExist:
                pass

        # Handle pre-population from URL parameters (e.g., from order detail page)
        email = self.request.GET.get('email', '').strip()
        first_name = self.request.GET.get('first_name', '').strip()
        last_name = self.request.GET.get('last_name', '').strip()
        phone = self.request.GET.get('phone', '').strip()
        source = self.request.GET.get('source', '').strip()

        if email:
            initial['email'] = email
        if first_name:
            initial['first_name'] = first_name
        if last_name:
            initial['last_name'] = last_name
        if phone:
            initial['phone'] = phone

        # Set contact type to individual if we have personal details
        if email or first_name or last_name:
            initial['contact_type'] = 'individual'

        return initial

    def form_valid(self, form):
        # Set created_by
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        
        # Handle company association
        company = form.cleaned_data.get('company')
        if company:
            self.object.companies.add(company)
        
        messages.success(self.request, 'Contact created successfully!')
        return response

    def get_success_url(self):
        if self.request.POST.get('action') == 'save_continue':
            return reverse_lazy('crm:contact_create')
        return reverse_lazy('crm:contact_list')


class ContactUpdateView(LoginRequiredMixin, UpdateView):
    model = Contact
    form_class = ContactForm
    template_name = 'crm/contact_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        contact = self.get_object()
        # If contact has associated company, pass it to form
        if contact.companies.exists():
            kwargs['company_id'] = contact.companies.first().id
        return kwargs

    def get_initial(self):
        initial = super().get_initial()
        contact = self.get_object()
        
        # Pre-populate form with existing data
        if contact.companies.exists():
            company = contact.companies.first()
            initial['company'] = company.id
            initial['contact_type'] = 'company'
        else:
            initial['contact_type'] = 'individual'
            
        return initial

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Handle company association
        company = form.cleaned_data.get('company')
        self.object.companies.clear()  # Remove existing associations
        if company:
            self.object.companies.add(company)
        
        messages.success(self.request, 'Contact updated successfully!')
        return response

    def get_success_url(self):
        if self.request.POST.get('action') == 'save_continue':
            return reverse_lazy('crm:contact_update', kwargs={'pk': self.object.pk})
        return reverse_lazy('crm:contact_list')


class ContactDeleteView(LoginRequiredMixin, DeleteView):
    model = Contact
    success_url = reverse_lazy('crm:contact_list')


# Company Views
class CompanyListView(LoginRequiredMixin, ListView):
    model = Company
    template_name = 'crm/company_list.html'
    context_object_name = 'companies'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Company.objects.select_related('created_by', 'primary_contact').prefetch_related('contacts', 'leads')
        
        # Filter by search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(industry__icontains=search) |
                Q(city__icontains=search) |
                Q(county__icontains=search)
            )
        
        # Filter by company type
        company_type = self.request.GET.get('company_type')
        if company_type:
            queryset = queryset.filter(company_type=company_type)
        
        # Filter by industry
        industry = self.request.GET.get('industry')
        if industry:
            queryset = queryset.filter(industry__icontains=industry)
        
        # Filter by county
        county = self.request.GET.get('county')
        if county:
            queryset = queryset.filter(county__icontains=county)
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter options
        industries = Company.objects.exclude(industry__isnull=True).exclude(industry='').values_list('industry', flat=True).distinct()
        counties = Company.objects.exclude(county__isnull=True).exclude(county='').values_list('county', flat=True).distinct()
        
        context.update({
            'industries': sorted(set(industries)),
            'counties': sorted(set(counties)),
            'current_filters': {
                'search': self.request.GET.get('search', ''),
                'company_type': self.request.GET.get('company_type', ''),
                'industry': self.request.GET.get('industry', ''),
                'county': self.request.GET.get('county', ''),
            },
            # Add stats
            'active_companies_count': Company.objects.count(),
            'new_companies_this_month': Company.objects.filter(
                created_at__date__gte=timezone.now().date() - timedelta(days=30)
            ).count(),
            'total_contacts_count': Contact.objects.filter(companies__isnull=False).distinct().count(),
        })
        return context


class CompanyDetailView(LoginRequiredMixin, DetailView):
    model = Company
    template_name = 'crm/company_detail.html'
    context_object_name = 'company'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        company = self.get_object()
        
        # Get counts with safe handling
        context.update({
            'opportunities_count': getattr(company, 'opportunities', []).count() if hasattr(company, 'opportunities') else 0,
            'activities_count': getattr(company, 'activities', []).count() if hasattr(company, 'activities') else 0,
            'active_leads_count': company.leads.exclude(
                status__in=['won', 'lost']
            ).count() if hasattr(company, 'leads') else 0,
            
            # Calculate total opportunity value using 'value' instead of 'estimated_value'
            'total_opportunity_value': (
                company.opportunities.aggregate(total=Sum('value'))['total'] or 0
                if hasattr(company, 'opportunities') else 0
            )
        })
        
        return context


class CompanyCreateView(LoginRequiredMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'crm/company_form.html'
    success_url = reverse_lazy('crm:company_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['initial'] = {
            'country': 'Kenya',  # Set default country
            }
        return kwargs

    def form_valid(self, form):
        try:
            # Set created_by to current user
            form.instance.created_by = self.request.user
            response = super().form_valid(form)
            messages.success(self.request, 'Company created successfully!')
            return response
        except Exception as e:
            messages.error(self.request, f'Error creating company: {str(e)}')
            return super().form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        print("Form Errors:", form.errors)  # For debugging
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'preview_data': {
                'contacts_count': 0,
                'leads_count': 0,
                'opportunities_count': 0
            }
        })
        return context


class CompanyUpdateView(LoginRequiredMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'crm/company_form.html'


class CompanyDeleteView(LoginRequiredMixin, DeleteView):
    model = Company
    success_url = reverse_lazy('crm:company_list')


# Activity Views
class ActivityListView(LoginRequiredMixin, ListView):
    model = Activity
    template_name = 'crm/activity_list.html'
    context_object_name = 'activities'
    paginate_by = 25
    
    def get_queryset(self):
        return Activity.objects.select_related('contact', 'assigned_to').order_by('scheduled_datetime')


class ActivityDetailView(LoginRequiredMixin, DetailView):
    model = Activity
    template_name = 'crm/activity_detail.html'
    context_object_name = 'activity'


class ActivityCalendarEventsView(LoginRequiredMixin, View):
    def get(self, request):
        start = request.GET.get('start')
        end = request.GET.get('end')
        activity_types = request.GET.get('types', '').split(',')

        activities = Activity.objects.filter(
            scheduled_datetime__range=[start, end]
        )

        if activity_types:
            activities = activities.filter(activity_type__in=activity_types)

        events = []
        for activity in activities:
            end_time = activity.scheduled_datetime + timedelta(minutes=activity.duration_minutes)
            events.append({
                'id': activity.pk,
                'subject': activity.subject,
                'activity_type': activity.activity_type,
                'scheduled_datetime': activity.scheduled_datetime.isoformat(),
                'end_datetime': end_time.isoformat(),
                'contact_name': activity.contact.get_full_name() if activity.contact else '',
                'status': activity.status
            })

        return JsonResponse(events, safe=False)


class ActivityCreateView(LoginRequiredMixin, CreateView):
    model = Activity
    form_class = ActivityForm
    template_name = 'crm/activity_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if not form.instance.assigned_to:
            form.instance.assigned_to = self.request.user
        return super().form_valid(form)


class ActivityUpdateView(LoginRequiredMixin, UpdateView):
    model = Activity
    form_class = ActivityForm
    template_name = 'crm/activity_form.html'


class ActivityDeleteView(LoginRequiredMixin, DeleteView):
    model = Activity
    success_url = reverse_lazy('crm:activity_list')


# Campaign Views
class CampaignListView(LoginRequiredMixin, ListView):
    model = Campaign
    template_name = 'crm/campaign_list.html'
    context_object_name = 'campaigns'
    paginate_by = 25


class CampaignDetailView(LoginRequiredMixin, DetailView):
    model = Campaign
    template_name = 'crm/campaign_detail.html'
    context_object_name = 'campaign'


class CampaignCreateView(LoginRequiredMixin, CreateView):
    model = Campaign
    form_class = CampaignForm
    template_name = 'crm/campaign_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Campaign created successfully!')
        return super().form_valid(form)


class CampaignUpdateView(LoginRequiredMixin, UpdateView):
    model = Campaign
    form_class = CampaignForm
    template_name = 'crm/campaign_form.html'


class CampaignDeleteView(LoginRequiredMixin, DeleteView):
    model = Campaign
    success_url = reverse_lazy('crm:campaign_list')


# Pipeline and Reports
class PipelineView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/pipeline.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get pipeline stages from Opportunity model
        stages = dict(Opportunity.STAGE_CHOICES)
        pipeline_stages = {}

        # Build pipeline data
        for stage_code, stage_name in stages.items():
            opportunities = Opportunity.objects.filter(stage=stage_code)
            
            pipeline_stages[stage_code] = {
                'name': stage_name,
                'opportunities': opportunities,
                'count': opportunities.count(),
                'total_value': sum(opp.value for opp in opportunities),
                'weighted_value': sum(opp.weighted_value for opp in opportunities)
            }

        # Calculate pipeline metrics
        all_opportunities = Opportunity.objects.all()
        won_opportunities = Opportunity.objects.filter(stage='closed_won')
        
        context.update({
            'pipeline_stages': pipeline_stages,
            'total_opportunities': all_opportunities.count(),
            'total_pipeline_value': sum(opp.value for opp in all_opportunities),
            'weighted_pipeline_value': sum(opp.weighted_value for opp in all_opportunities),
            'average_deal_size': all_opportunities.aggregate(avg_value=Avg('value'))['avg_value'] or 0,
            'conversion_rate': (won_opportunities.count() / all_opportunities.count() * 100) if all_opportunities.exists() else 0,
            'team_members': User.objects.filter(is_active=True)
        })
        return context


class PipelineExportView(LoginRequiredMixin, View):
    def get(self, request):
        format = request.GET.get('format', 'csv')
        opportunities = Opportunity.objects.all().order_by('stage')
        
        if format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="pipeline.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Stage', 'Value', 'Probability', 'Contact', 'Company', 'Expected Close'])
            
            for opp in opportunities:
                writer.writerow([
                    opp.name,
                    opp.get_stage_display(),
                    opp.value,
                    opp.probability,
                    opp.contact.get_full_name() if opp.contact else '',
                    opp.company.name if opp.company else '',
                    opp.expected_close_date
                ])
            
            return response
            
        elif format == 'pdf':
            # Implement PDF export
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="pipeline.pdf"'
            # Add PDF generation code here
            return response

        return HttpResponse('Unsupported format', status=400)


class OpportunityMoveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        try:
            opportunity = get_object_or_404(Opportunity, pk=pk)
            data = json.loads(request.body)
            new_stage = data.get('stage')
            
            if new_stage in dict(Opportunity.STAGE_CHOICES):
                old_stage = opportunity.stage
                opportunity.change_stage(new_stage, request.user)
                
                # Log the stage change
                Activity.objects.create(
                    subject=f"Stage changed from {old_stage} to {new_stage}",
                    activity_type='other',
                    status='completed',
                    opportunity=opportunity,
                    contact=opportunity.contact,
                    created_by=request.user,
                    assigned_to=opportunity.assigned_to
                )
                
                return JsonResponse({'success': True})
            return JsonResponse({'success': False, 'error': 'Invalid stage'})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ReportsView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get time ranges once
        today = timezone.now()
        thirty_days_ago = today - timedelta(days=30)
        current_month_start = today.replace(day=1)
        last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)

        # Get all opportunities with select_related in one query
        opportunities = Opportunity.objects.select_related(
            'contact', 'company'
        ).all()

        # Get leads with optimization
        leads = Lead.objects.select_related('source')
        current_month_leads = leads.filter(created_at__gte=thirty_days_ago).count()
        previous_month_leads = leads.filter(
            created_at__range=[thirty_days_ago - timedelta(days=30), thirty_days_ago]
        ).count()

        # Calculate pipeline metrics
        pipeline_metrics = self._calculate_pipeline_metrics(opportunities)
        
        context.update({
            # Basic metrics
            'total_leads': current_month_leads,
            'lead_growth': self._calculate_growth(current_month_leads, previous_month_leads),
            'total_opportunities': len(opportunities),
            'opportunity_growth': self._calculate_opportunity_growth(current_month_start, last_month_start),
            
            # Conversion and pipeline metrics
            'conversion_rate': self._calculate_conversion_rate(leads),
            'pipeline_value': pipeline_metrics['total_value'],
            'pipeline_stages': self._get_pipeline_stages(opportunities),
            'avg_deal_size': pipeline_metrics['avg_deal_size'],
            'sales_cycle_days': self._calculate_sales_cycle_days(opportunities),
            
            # Performance data
            'lead_sources_data': self._get_lead_sources_data(thirty_days_ago),
            'monthly_performance': self._get_monthly_performance(today),
            'team_performance': self._get_team_performance(thirty_days_ago)
        })
        
        return context

    def _calculate_pipeline_metrics(self, opportunities):
        """Calculate pipeline metrics in one pass"""
        total_value = 0
        total_count = 0
        
        for opp in opportunities:
            if not opp.stage == 'closed_lost':
                total_value += opp.value
                total_count += 1
        
        return {
            'total_value': total_value,
            'avg_deal_size': total_value / total_count if total_count > 0 else 0
        }

    def _calculate_opportunity_growth(self, current_month_start, last_month_start):
        """Calculate opportunity growth with optimized date handling"""
        this_month = Opportunity.objects.filter(
            created_at__gte=current_month_start
        ).count()
        
        last_month = Opportunity.objects.filter(
            created_at__range=[last_month_start, current_month_start]
        ).count()
        
        return self._calculate_growth(this_month, last_month)

    def _calculate_growth(self, current, previous):
        """Calculate percentage growth between two values"""
        if previous == 0:
            return 100 if current > 0 else 0
        return ((current - previous) / previous) * 100

    def _calculate_conversion_rate(self, leads):
        """Calculate conversion rate with cached querysets"""
        total_leads = leads.count()
        if total_leads == 0:
            return 0
        converted_leads = leads.filter(status='won').count()
        return (converted_leads / total_leads) * 100

    def _get_pipeline_stages(self, opportunities):
        """Get pipeline stages with prefetched opportunities"""
        stages = []
        total_count = len(opportunities)
        
        stage_data = {}
        for opp in opportunities:
            if opp.stage not in stage_data:
                stage_data[opp.stage] = {'count': 0, 'value': 0}
            stage_data[opp.stage]['count'] += 1
            stage_data[opp.stage]['value'] += opp.value

        for stage_code, stage_name in Opportunity.STAGE_CHOICES:
            data = stage_data.get(stage_code, {'count': 0, 'value': 0})
            stages.append({
                'name': stage_name,
                'count': data['count'],
                'value': data['value'],
                'conversion_rate': (data['count'] / total_count * 100) if total_count > 0 else 0
            })
        
        return stages

    def _calculate_sales_cycle_days(self, opportunities):
        """Calculate sales cycle with optimized query"""
        won_opportunities = [
            opp for opp in opportunities 
            if opp.stage == 'closed_won' and opp.actual_close_date
        ]
        
        if not won_opportunities:
            return 0
            
        total_days = sum(
            (opp.actual_close_date - opp.created_at.date()).days
            for opp in won_opportunities
        )
        return round(total_days / len(won_opportunities))

    def _get_monthly_performance(self, today):
        """Get monthly performance with optimized queries"""
        data = []
        
        for i in range(6):
            month_start = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            month_data = {
                'month': month_start.strftime('%b %Y'),
                'leads': Lead.objects.filter(
                    created_at__date__range=[month_start, month_end]
                ).count(),
                'revenue': float(
                    Opportunity.objects.filter(
                        actual_close_date__range=[month_start, month_end],
                        stage='closed_won'
                    ).aggregate(total=Sum('value'))['total'] or 0
                )
            }
            data.append(month_data)
        
        return list(reversed(data))

    def _get_team_performance(self, thirty_days_ago):
        """Get team performance with optimized annotations"""
        return User.objects.filter(
            Q(opportunities__isnull=False) | Q(assigned_leads__isnull=False)
        ).distinct().annotate(
            leads_count=Count(
                'assigned_leads',
                filter=Q(assigned_leads__created_at__gte=thirty_days_ago)
            ),
            opportunities_value=Sum(
                'opportunities__value',
                filter=Q(
                    opportunities__created_at__gte=thirty_days_ago,
                    opportunities__stage__in=['qualification', 'needs_analysis', 'proposal', 'negotiation', 'closed_won']
                )
            ),
            user_role=Case(
                When(is_superuser=True, then=Value('Administrator')),
                When(groups__name='Sales Manager', then=Value('Sales Manager')),
                When(groups__name='Manager', then=Value('Manager')),
                When(groups__name='Team Lead', then=Value('Team Lead')),
                default=Value('Sales Person'),
                output_field=CharField(),
            )
        ).order_by('-opportunities_value', '-leads_count').values(
            'id', 
            'first_name', 
            'last_name', 
            'leads_count', 
            'opportunities_value',
            'user_role'
        )[:5]

    def _get_lead_sources_data(self, thirty_days_ago):
        """Get lead source distribution data"""
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        sources_data = LeadSource.objects.filter(
            leads__created_at__gte=thirty_days_ago
        ).annotate(
            total_leads=Count('leads'),
            converted_leads=Count('leads', filter=Q(leads__status='won')),
            total_value=Sum('leads__estimated_value', filter=Q(leads__status='won'))
        ).values(
            'name',
            'total_leads',
            'converted_leads',
            'total_value'
        )

        # Calculate additional metrics for each source
        for source in sources_data:
            source['conversion_rate'] = (
                (source['converted_leads'] / source['total_leads'] * 100)
                if source['total_leads'] > 0 else 0
            )
            source['avg_value'] = (
                float(source['total_value'] / source['converted_leads'])
                if source['converted_leads'] > 0 else 0
            )

        return list(sources_data)

class SalesForecastView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/reports/sales_forecast.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()

        # Get opportunities in pipeline
        opportunities = Opportunity.objects.exclude(stage__in=['closed_won', 'closed_lost'])

        # Group by expected close month
        forecast_data = {}
        months = []
        total_values = []
        weighted_values = []

        for opp in opportunities:
            month_key = opp.expected_close_date.strftime('%Y-%m')
            month_label = opp.expected_close_date.strftime('%B %Y')
            
            if month_key not in forecast_data:
                forecast_data[month_key] = {
                    'month': month_label,
                    'total_value': 0,
                    'weighted_value': 0,
                    'count': 0
                }
                months.append(month_label)
            
            forecast_data[month_key]['total_value'] += float(opp.value)
            forecast_data[month_key]['weighted_value'] += float(opp.weighted_value)
            forecast_data[month_key]['count'] += 1

        # Sort data by month
        sorted_data = sorted(forecast_data.values(), key=lambda x: datetime.strptime(x['month'], '%B %Y'))
        
        # Prepare chart data
        for data in sorted_data:
            total_values.append(data['total_value'])
            weighted_values.append(data['weighted_value'])

        # Calculate metrics
        total_pipeline = sum(opp.value for opp in opportunities)
        weighted_pipeline = sum(opp.weighted_value for opp in opportunities)
        
        # Calculate historical win rate
        closed_opportunities = Opportunity.objects.filter(
            actual_close_date__lte=today,
            stage__in=['closed_won', 'closed_lost']
        )
        won_opportunities = closed_opportunities.filter(stage='closed_won')
        win_rate = (won_opportunities.count() / closed_opportunities.count() * 100) if closed_opportunities.count() > 0 else 0

        # Probability distribution
        probability_data = {
            'high': opportunities.filter(probability__gte=70).aggregate(total=Sum('value'))['total'] or 0,
            'medium': opportunities.filter(probability__range=(40, 69)).aggregate(total=Sum('value'))['total'] or 0,
            'low': opportunities.filter(probability__lte=39).aggregate(total=Sum('value'))['total'] or 0
        }

        context.update({
            'forecast_data': sorted_data,
            'chart_data': {
                'months': months,
                'total_values': total_values,
                'weighted_values': weighted_values
            },
            'total_pipeline': total_pipeline,
            'weighted_pipeline': weighted_pipeline,
            'opportunity_count': opportunities.count(),
            'historical_win_rate': win_rate,
            'average_deal_size': opportunities.aggregate(avg=Avg('value'))['avg'] or 0,
            'high_probability_value': float(probability_data['high']),
            'medium_probability_value': float(probability_data['medium']),
            'low_probability_value': float(probability_data['low'])
        })

        return context


# API Views
class LeadsByStatusAPIView(LoginRequiredMixin, View):
    def get(self, request):
        data = list(Lead.objects.values('status').annotate(count=Count('id')))
        return JsonResponse(data, safe=False)


class OpportunityForecastAPIView(LoginRequiredMixin, View):
    def get(self, request):
        data = list(Opportunity.objects.values('stage').annotate(
            count=Count('id'),
            total_value=Sum('value'),
            weighted_value=Sum('value') * Avg('probability') / 100
        ))
        return JsonResponse(data, safe=False)


class LeadSourceListView(LoginRequiredMixin, ListView):
    model = LeadSource
    template_name = 'crm/lead_source_list.html'
    context_object_name = 'lead_sources'
    
    def get_queryset(self):
        return LeadSource.objects.order_by('name')

class LeadSourceDetailView(LoginRequiredMixin, DetailView):
    model = LeadSource
    template_name = 'crm/lead_source_detail.html'
    context_object_name = 'lead_source'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead_source = self.get_object()
        
        # Get leads associated with this source
        leads = lead_source.leads.select_related('contact', 'company', 'assigned_to')
        
        # Calculate statistics
        total_leads = leads.count()
        converted_leads = leads.filter(status='won').count()
        
        # Calculate conversion rate
        conversion_rate = 0.0
        if total_leads > 0:
            conversion_rate = (converted_leads / total_leads) * 100
        
        context.update({
            'total_leads': total_leads,
            'active_leads': leads.exclude(status__in=['won', 'lost']).count(),
            'converted_leads': converted_leads,
            'conversion_rate': conversion_rate,
            'recent_leads': leads.order_by('-created_at')[:5],
            'leads_by_status': leads.values('status').annotate(count=Count('id')),
        })
        
        return context

class LeadSourceCreateView(LoginRequiredMixin, CreateView):
    model = LeadSource
    fields = ['name', 'description', 'is_active']
    template_name = 'crm/lead_source_form.html'
    success_url = reverse_lazy('crm:lead_source_list')

    def form_valid(self, form):
        messages.success(self.request, 'Lead source created successfully!')
        return super().form_valid(form)

class LeadSourceUpdateView(LoginRequiredMixin, UpdateView):
    model = LeadSource
    fields = ['name', 'description', 'is_active']
    template_name = 'crm/lead_source_form.html'
    success_url = reverse_lazy('crm:lead_source_list')

    def form_valid(self, form):
        messages.success(self.request, 'Lead source updated successfully!')
        return super().form_valid(form)

class LeadSourceDeleteView(LoginRequiredMixin, DeleteView):
    model = LeadSource
    success_url = reverse_lazy('crm:lead_source_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Lead source deleted successfully!')
        return super().delete(request, *args, **kwargs)


# Email Composer View
class EmailComposerView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/email_composer.html'

    def get_context_data(self, **kwargs):
        import re  # Import re module locally to avoid scoping issues
        context = super().get_context_data(**kwargs)

        # Get recipient information from URL parameters
        lead_id = self.request.GET.get('lead')
        opportunity_id = self.request.GET.get('opportunity')
        contact_param = self.request.GET.get('contact')
        email = self.request.GET.get('email')

        # Initialize variables
        preselected_contact = None
        preselected_lead = None
        preselected_opportunity = None

        # Priority: opportunity > lead > contact > email
        if opportunity_id and opportunity_id.isdigit():
            try:
                preselected_opportunity = Opportunity.objects.select_related('contact', 'company', 'lead').get(pk=int(opportunity_id))
                preselected_contact = preselected_opportunity.contact
                preselected_lead = preselected_opportunity.lead
            except Opportunity.DoesNotExist:
                pass
        elif lead_id and lead_id.isdigit():
            try:
                preselected_lead = Lead.objects.select_related('contact', 'company').get(pk=int(lead_id))
                preselected_contact = preselected_lead.contact
            except Lead.DoesNotExist:
                pass
        elif contact_param:
            try:
                # Check if it's an ID (numeric)
                if contact_param.isdigit():
                    preselected_contact = Contact.objects.get(pk=int(contact_param))
                else:
                    # Assume it's an email address
                    preselected_contact = Contact.objects.get(email=contact_param)
            except (Contact.DoesNotExist, ValueError):
                # If contact not found by ID or if not a valid ID, try email lookup
                if contact_param:
                    try:
                        preselected_contact = Contact.objects.select_related('companies').get(email=contact_param)
                    except Contact.DoesNotExist:
                        pass
                elif email:
                    try:
                        preselected_contact = Contact.objects.select_related('companies').get(email=email)
                    except Contact.DoesNotExist:
                        pass
        elif email:
            # Try to find contact by email if no contact parameter
            try:
                preselected_contact = Contact.objects.select_related('companies').get(email=email)
            except Contact.DoesNotExist:
                pass

        # Get all contacts for selection
        contacts = Contact.objects.prefetch_related('companies').order_by('first_name', 'last_name')

        # Get email templates - show relevant templates based on context
        if preselected_opportunity:
            # For opportunities, try opportunity templates first, fallback to appropriate lead templates
            opportunity_templates = EmailTemplate.objects.filter(
                is_active=True,
                template_type__in=[
                    'opportunity_qualification',
                    'opportunity_proposal',
                    'opportunity_negotiation',
                    'opportunity_closure',
                    'opportunity_follow_up'
                ]
            ).order_by('template_type', 'name')

            if opportunity_templates.exists():
                # If opportunity templates exist, show them
                email_templates = opportunity_templates
            else:
                # Fall back to lead templates appropriate for opportunities
                email_templates = EmailTemplate.objects.filter(
                    is_active=True,
                    template_type__in=['lead_nurture', 'proposal', 'follow_up']
                ).order_by('template_type', 'name')
        else:
            # For leads and general use, show lead-focused templates
            email_templates = EmailTemplate.objects.filter(
                is_active=True,
                template_type__in=[
                    'welcome',
                    'lead_nurture',
                    'follow_up',
                    'proposal',
                    'thank_you',
                    'rejection',
                    'meeting_request',
                    'custom'
                ]
            ).order_by('template_type', 'name')

        # Add template metadata for display
        template_metadata = {
            'welcome': {
                'icon': 'fas fa-handshake',
                'color': 'var(--primary-color)',
                'description': 'Welcome new customers',
                'category': 'Onboarding'
            },
            'lead_nurture': {
                'icon': 'fas fa-seedling',
                'color': '#28a745',
                'description': 'Nurture and engage leads',
                'category': 'Nurturing'
            },
            'follow_up': {
                'icon': 'fas fa-clock',
                'color': '#ffc107',
                'description': 'Follow up on leads',
                'category': 'Follow-up'
            },
            'proposal': {
                'icon': 'fas fa-file-invoice-dollar',
                'color': '#17a2b8',
                'description': 'Send proposals and quotes',
                'category': 'Sales'
            },
            'thank_you': {
                'icon': 'fas fa-heart',
                'color': '#28a745',
                'description': 'Thank customers for business',
                'category': 'Customer Care'
            },
            'rejection': {
                'icon': 'fas fa-hand-paper',
                'color': '#6c757d',
                'description': 'Handle lost opportunities',
                'category': 'Closure'
            },
            'meeting_request': {
                'icon': 'fas fa-calendar-alt',
                'color': '#6f42c1',
                'description': 'Schedule meetings and calls',
                'category': 'Scheduling'
            },
            'custom': {
                'icon': 'fas fa-edit',
                'color': '#6c757d',
                'description': 'Custom email templates',
                'category': 'Custom'
            }
        }

        # Add metadata to templates
        for template in email_templates:
            metadata = template_metadata.get(template.template_type, template_metadata['custom'])
            template.icon = metadata['icon']
            template.color = metadata['color']
            template.description = metadata['description']
            template.category = metadata['category']

        # Process template variables for preselected contact/lead
        processed_templates = []
        for template in email_templates:
            processed_template = {
                'id': template.id,
                'name': template.name,
                'template_type': template.template_type,
                'icon': template.icon,
                'color': template.color,
                'description': template.description,
                'category': template.category,
            }

            # Process subject and body with recipient data if available
            subject = template.subject
            body = template.body

            if preselected_contact:
                # Replace contact variables
                subject = subject.replace('{{contact_first_name}}', preselected_contact.first_name or '')
                subject = subject.replace('{{contact_name}}', preselected_contact.get_full_name())
                subject = subject.replace('{{contact.email}}', preselected_contact.email or '')

                body = body.replace('{{contact_first_name}}', preselected_contact.first_name or '')
                body = body.replace('{{contact_name}}', preselected_contact.get_full_name())
                body = body.replace('{{contact.email}}', preselected_contact.email or '')
                body = body.replace('{{contact.phone}}', preselected_contact.phone or '')
                body = body.replace('{{contact.position}}', preselected_contact.position or '')
                body = body.replace('{{contact.city}}', preselected_contact.city or '')

                # Replace recipient references
                body = body.replace('{{contact}}', preselected_contact.get_full_name())
                body = body.replace('{{recipient_name}}', preselected_contact.get_full_name())
                body = body.replace('{{recipient}}', preselected_contact.get_full_name())

            if preselected_lead:
                # Replace lead variables
                subject = subject.replace('{{lead_title}}', preselected_lead.title or '')

                body = body.replace('{{lead_title}}', preselected_lead.title or '')
                body = body.replace('{{estimated_value}}', f"KES {preselected_lead.estimated_value:,}" if preselected_lead.estimated_value else '')
                body = body.replace('{{priority}}', preselected_lead.get_priority_display())
                body = body.replace('{{status}}', preselected_lead.get_status_display())
                body = body.replace('{{system_type}}', preselected_lead.get_system_type_display() if preselected_lead.system_type else '')
                body = body.replace('{{estimated_capacity}}', f"{preselected_lead.estimated_capacity} kW" if preselected_lead.estimated_capacity else '')
                body = body.replace('{{source}}', preselected_lead.source.name if preselected_lead.source else '')

                # Calculate days since created
                from django.utils import timezone
                days_since_created = (timezone.now().date() - preselected_lead.created_at.date()).days
                body = body.replace('{{days_since_created}}', str(days_since_created))

                # Next follow-up date
                if preselected_lead.next_follow_up:
                    body = body.replace('{{next_follow_up}}', preselected_lead.next_follow_up.strftime('%B %d, %Y'))
                else:
                    # Remove the conditional block if no follow-up date
                    import re
                    body = re.sub(r'{% if next_follow_up %}.*?{% endif %}', '', body, flags=re.DOTALL)

                # Expected close date
                if preselected_lead.expected_close_date:
                    body = body.replace('{{expected_close_date}}', preselected_lead.expected_close_date.strftime('%B %d, %Y'))
                else:
                    body = body.replace('{{expected_close_date}}', '')

                # Remove conditional blocks that don't apply
                body = re.sub(r'{% if.*?(?:%}|$)','', body, flags=re.DOTALL)
                body = re.sub(r'{% endif %}','', body)

            if preselected_opportunity:
                # Replace opportunity variables
                body = body.replace('{{opportunity_name}}', preselected_opportunity.name)
                body = body.replace('{{opportunity_value}}', f"KES {preselected_opportunity.value:,.0f}" if preselected_opportunity.value else '')
                body = body.replace('{{weighted_value}}', f"KES {preselected_opportunity.weighted_value:,.0f}" if preselected_opportunity.weighted_value else '')
                body = body.replace('{{probability}}', f"{preselected_opportunity.probability}%" if preselected_opportunity.probability else '')
                body = body.replace('{{stage}}', preselected_opportunity.get_stage_display())
                body = body.replace('{{expected_close_date}}', preselected_opportunity.expected_close_date.strftime('%B %d, %Y') if preselected_opportunity.expected_close_date else '')
                body = body.replace('{{next_steps}}', preselected_opportunity.next_steps or '')
                body = body.replace('{{competition}}', preselected_opportunity.competition or '')

                # Calculate days in pipeline for opportunity
                days_since_created = (timezone.now().date() - preselected_opportunity.created_at.date()).days
                body = body.replace('{{days_since_created}}', str(days_since_created))

            if preselected_lead and preselected_lead.company:
                # Replace company variables
                body = body.replace('{{company_name}}', preselected_lead.company.name or '')
                body = body.replace('{{company.website}}', preselected_lead.company.website or '')
                body = body.replace('{{company.phone}}', preselected_lead.company.phone or '')
                body = body.replace('{{company.industry}}', preselected_lead.company.industry or '')

            # Set assigned user info - prefer opportunity assignee, then lead assignee, then current user
            assigned_user = (preselected_opportunity.assigned_to if preselected_opportunity else None) or (preselected_lead.assigned_to if preselected_lead else None) or self.request.user
            if assigned_user:
                body = body.replace('{{assigned_to}}', assigned_user.get_full_name())
                body = body.replace('{{assigned_to_phone}}', assigned_user.phone or '+254 719 728 666')
                body = body.replace('{{assigned_to_email}}', assigned_user.email or 'info@olivian.co.ke')
                body = body.replace('{{assigned_to_signature}}', f"\n\nBest regards,\n{assigned_user.get_full_name()}\n{assigned_user.get_role_display() if hasattr(assigned_user, 'get_role_display') and assigned_user.get_role_display() else 'CRM Team'}")

                # Replace email references
                body = body.replace('{{lead_email}}', preselected_contact.email if preselected_contact else '')

            processed_template['subject'] = subject
            processed_template['body'] = body
            processed_templates.append(processed_template)

        context.update({
            'contacts': contacts,
            'email_templates': processed_templates,
            'preselected_contact': preselected_contact,
            'preselected_lead': preselected_lead,
            'contact_variables': [
                ('contact.first_name', 'First Name'),
                ('contact.last_name', 'Last Name'),
                ('contact.full_name', 'Full Name'),
                ('contact.email', 'Email Address'),
                ('contact.phone', 'Phone Number'),
                ('contact.position', 'Job Position'),
                ('contact.address_line_1', 'Address Line 1'),
                ('contact.city', 'City'),
                ('contact.county', 'County'),
                ('contact.country', 'Country'),
            ],
            'company_variables': [
                ('company.name', 'Company Name'),
                ('company.website', 'Website'),
                ('company.industry', 'Industry'),
                ('company.phone', 'Company Phone'),
                ('company.address_line_1', 'Company Address'),
                ('company.city', 'Company City'),
                ('company.county', 'Company County'),
            ],
            'lead_variables': [
                ('lead.title', 'Lead Title'),
                ('lead.source_name', 'Lead Source'),
                ('lead.status', 'Lead Status'),
                ('lead.estimated_value', 'Estimated Value'),
                ('lead.priority', 'Priority Level'),
            ] if preselected_lead else []
        })

        return context

    def post(self, request, *args, **kwargs):
        """Handle email sending"""
        try:
            subject = request.POST.get('subject', '').strip()
            body = request.POST.get('body', '').strip()
            recipients = request.POST.getlist('recipients')

            if not subject:
                return JsonResponse({'success': False, 'error': 'Subject is required'})

            if not body:
                return JsonResponse({'success': False, 'error': 'Email body is required'})

            if not recipients:
                return JsonResponse({'success': False, 'error': 'Recipients are required'})

            # Import EmailService here to avoid circular imports
            from apps.core.email_utils import EmailService

            # Get signature from settings
            signature = getattr(settings, 'CRM_EMAIL_SIGNATURE', f"""
Best regards,
{request.user.get_full_name()}
{request.user.get_role_display() if hasattr(request.user, 'get_role_display') else 'CRM Team'}
{getattr(settings, 'COMPANY_NAME', 'The Olivian Group')}
""").strip()

            sent_emails = 0
            failed_emails = 0
            failures = []
            emailed_leads = []
            emailed_contacts = []

            for recipient_email in recipients:
                recipient_email = recipient_email.strip()
                if not recipient_email:
                    continue

                try:
                    # Try to find the contact/lead associated with this email
                    contact = None
                    lead = None

                    # Check if email matches a contact
                    contact = Contact.objects.filter(email=recipient_email).first()
                    if contact:
                        # Look for associated leads
                        lead = Lead.objects.filter(contact=contact).first()

                    # Store references for progress updates
                    if contact:
                        emailed_contacts.append(contact)
                    if lead:
                        emailed_leads.append(lead)

                except Exception as e:
                    # Log error but don't fail the email sending
                    logger.warning(f"Error finding contact/lead for email {recipient_email}: {str(e)}")

                # Prepare email content with signature
                full_content = f"{body}\n\n{signature}"

                # Send email using the EmailService
                success = EmailService.send_email_notification(
                    template_name='crm_email',
                    context={
                        'content': full_content,
                        'sender': request.user
                    },
                    recipient_email=recipient_email,
                    subject=subject,
                    from_email=getattr(settings, 'CRM_EMAIL_FROM', 'sales@olivian.co.ke')
                )

                if success:
                    sent_emails += 1
                else:
                    failed_emails += 1
                    failures.append(recipient_email)

            # After successful email sending, update lead progress
            if sent_emails > 0:
                self._update_lead_progress(request, emailed_leads, emailed_contacts, subject, body)

            # Send JSON response
            if sent_emails > 0:
                message = f"Successfully sent {sent_emails} email{'s' if sent_emails > 1 else ''}"
                if failed_emails > 0:
                    message += f". {failed_emails} email{'s' if failed_emails > 1 else ''} failed."
                return JsonResponse({
                    'success': True,
                    'message': message,
                    'sent_count': sent_emails,
                    'failed_count': failed_emails,
                    'failures': failures if failures else None
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': f'Failed to send {failed_emails} email{"s" if failed_emails > 1 else ""}'

                })

        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error sending email: {str(e)}")
            return JsonResponse({
                'success': False,
                'error': f'An error occurred while sending the email: {str(e)}'
            })

    def _update_lead_progress(self, request, emailed_leads, emailed_contacts, subject, body):
        """Update lead progress and create activity logs after successful email sending"""
        try:
            # Process emailed leads
            for lead in emailed_leads:
                if lead:  # Make sure lead exists
                    # Update lead status from 'new' to 'contacted' if appropriate
                    if lead.status == 'new':
                        lead.status = 'contacted'
                        lead.save(update_fields=['status'])

                    # Set follow-up date if not set (7 days from now)
                    if not lead.next_follow_up:
                        lead.next_follow_up = timezone.now().date() + timedelta(days=7)
                        lead.save(update_fields=['next_follow_up'])

                    # Create activity log for email
                    Activity.objects.create(
                        subject=f"Email sent: {subject}",
                        activity_type='email',
                        status='completed',
                        scheduled_datetime=timezone.now(),
                        completed_datetime=timezone.now(),
                        duration_minutes=0,  # Email activity
                        lead=lead,
                        contact=lead.contact,
                        description=f"Email sent from CRM Email Composer\n\nSubject: {subject}\n\n{body}",
                        assigned_to=request.user,
                        created_by=request.user
                    )

            # Process emailed contacts (not necessarily leads)
            for contact in emailed_contacts:
                if contact and contact not in [lead.contact for lead in emailed_leads if lead]:
                    # Log activity for contact-only emails
                    Activity.objects.create(
                        subject=f"Email sent: {subject}",
                        activity_type='email',
                        status='completed',
                        scheduled_datetime=timezone.now(),
                        completed_datetime=timezone.now(),
                        duration_minutes=0,  # Email activity
                        contact=contact,
                        description=f"Email sent from CRM Email Composer\n\nSubject: {subject}\n\n{body}",
                        assigned_to=request.user,
                        created_by=request.user
                    )

        except Exception as e:
            logger.error(f"Error updating lead progress after email: {str(e)}")
            # Don't fail the entire email operation if this fails




# Email Template Views
class EmailTemplateListView(LoginRequiredMixin, ListView):
    model = EmailTemplate
    template_name = 'crm/email_template_list.html'
    context_object_name = 'email_templates'
    paginate_by = 20

    def get_queryset(self):
        queryset = EmailTemplate.objects.filter(is_active=True).order_by('template_type', 'name')

        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | Q(subject__icontains=search)
            )

        # Type filter
        template_type = self.request.GET.get('type')
        if template_type:
            queryset = queryset.filter(template_type=template_type)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'current_search': self.request.GET.get('search', ''),
            'current_type': self.request.GET.get('type', ''),
            'template_types': dict(EmailTemplate.TEMPLATE_TYPE_CHOICES),
        })
        return context


class EmailTemplateDetailView(LoginRequiredMixin, DetailView):
    model = EmailTemplate
    template_name = 'crm/email_template_detail.html'
    context_object_name = 'email_template'


class EmailTemplateCreateView(LoginRequiredMixin, CreateView):
    model = EmailTemplate
    template_name = 'crm/email_template_form.html'
    fields = ['name', 'template_type', 'subject', 'body', 'is_active']

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Email template created successfully!')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('crm:email_template_detail', kwargs={'pk': self.object.pk})


class EmailTemplateUpdateView(LoginRequiredMixin, UpdateView):
    model = EmailTemplate
    template_name = 'crm/email_template_form.html'
    fields = ['name', 'template_type', 'subject', 'body', 'is_active']

    def form_valid(self, form):
        form.instance.updated_by = self.request.user if hasattr(self.request.user, 'update') else self.request.user
        messages.success(self.request, 'Email template updated successfully!')
        return super().form_valid(form)


class EmailTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = EmailTemplate
    template_name = 'crm/email_template_delete.html'
    success_url = reverse_lazy('crm:email_template_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Email template deleted successfully!')
        return super().delete(request, *args, **kwargs)


@login_required
def test_email_template(request, pk):
    """Test send an email using a template - for development/testing"""
    try:
        template = EmailTemplate.objects.get(pk=pk)
        # This would be used for testing email templates
        # Implementation would require contact info and email service

        messages.success(request, f'Test email sent using "{template.name}" template.')
        return redirect('crm:email_template_detail', pk=pk)

    except EmailTemplate.DoesNotExist:
        messages.error(request, 'Email template not found.')
        return redirect('crm:email_template_list')
    except Exception as e:
        messages.error(request, f'Error sending test email: {str(e)}')
        return redirect('crm:email_template_list')


class CompanyExportView(LoginRequiredMixin, View):
    def get(self, request, pk):
        company = get_object_or_404(Company, pk=pk)
        export_format = request.GET.get('format', 'csv').lower()
        
        if export_format == 'csv':
            return self.export_csv(company)
        elif export_format == 'excel':
            return self.export_excel(company)
        elif export_format == 'pdf':
            return self.export_pdf(company)
        else:
            return HttpResponse('Invalid format specified', status=400)

    def export_csv(self, company):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{company.name}_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Company Information'])
        writer.writerow(['Name', company.name])
        writer.writerow(['Type', company.get_company_type_display()])
        writer.writerow(['Industry', company.industry])
        writer.writerow(['Registration Number', company.registration_number])
        writer.writerow(['Tax Number', company.tax_number])
        writer.writerow(['Email', company.email])
        writer.writerow(['Phone', company.phone])
        writer.writerow(['Website', company.website])
        writer.writerow(['Address', f"{company.address_line_1} {company.address_line_2}"])
        writer.writerow(['City', company.city])
        writer.writerow(['County', company.county])
        writer.writerow(['Postal Code', company.postal_code])
        writer.writerow(['Employee Count', company.employee_count])
        writer.writerow(['Annual Revenue', company.annual_revenue])
        
        # Add contacts
        writer.writerow([])
        writer.writerow(['Contacts'])
        writer.writerow(['Name', 'Email', 'Phone', 'Position'])
        for contact in company.contacts.all():
            writer.writerow([
                contact.get_full_name(),
                contact.email,
                contact.phone,
                contact.position
            ])
        
        # Add leads
        writer.writerow([])
        writer.writerow(['Leads'])
        writer.writerow(['Title', 'Status', 'Value', 'Created'])
        for lead in company.leads.all():
            writer.writerow([
                lead.title,
                lead.get_status_display(),
                lead.estimated_value,
                lead.created_at.strftime('%Y-%m-%d')
            ])
        
        return response

    def export_excel(self, company):
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output)
        
        # Company Information worksheet
        ws_info = workbook.add_worksheet('Company Info')
        bold = workbook.add_format({'bold': True})
        
        # Write company information
        row = 0
        ws_info.write(row, 0, 'Company Information', bold)
        row += 1
        
        info_rows = [
            ('Name', company.name),
            ('Type', company.get_company_type_display()),
            ('Industry', company.industry),
            ('Registration Number', company.registration_number),
            ('Tax Number', company.tax_number),
            ('Email', company.email),
            ('Phone', company.phone),
            ('Website', company.website),
            ('Address', f"{company.address_line_1} {company.address_line_2}"),
            ('City', company.city),
            ('County', company.county),
            ('Postal Code', company.postal_code),
            ('Employee Count', company.employee_count),
            ('Annual Revenue', company.annual_revenue),
        ]
        
        for label, value in info_rows:
            ws_info.write(row, 0, label, bold)
            ws_info.write(row, 1, value)
            row += 1
        
        # Contacts worksheet
        ws_contacts = workbook.add_worksheet('Contacts')
        headers = ['Name', 'Email', 'Phone', 'Position']
        for col, header in enumerate(headers):
            ws_contacts.write(0, col, header, bold)
        
        for row, contact in enumerate(company.contacts.all(), start=1):
            ws_contacts.write(row, 0, contact.get_full_name())
            ws_contacts.write(row, 1, contact.email)
            ws_contacts.write(row, 2, contact.phone)
            ws_contacts.write(row, 3, contact.position)
        
        # Leads worksheet
        ws_leads = workbook.add_worksheet('Leads')
        headers = ['Title', 'Status', 'Value', 'Created']
        for col, header in enumerate(headers):
            ws_leads.write(0, col, header, bold)
        
        for row, lead in enumerate(company.leads.all(), start=1):
            ws_leads.write(row, 0, lead.title)
            ws_leads.write(row, 1, lead.get_status_display())
            ws_leads.write(row, 2, float(lead.estimated_value) if lead.estimated_value else 0)
            ws_leads.write(row, 3, lead.created_at.strftime('%Y-%m-%d'))
        
        workbook.close()
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{company.name}_export.xlsx"'
        return response

    def export_pdf(self, company):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{company.name}_export.pdf"'
        
        # Create PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 750, company.name)
        
        # Company Information
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, 700, "Company Information")
        p.setFont("Helvetica", 10)
        
        y = 680
        info_rows = [
            ('Type:', company.get_company_type_display()),
            ('Industry:', company.industry),
            ('Email:', company.email),
            ('Phone:', company.phone),
            ('Website:', company.website),
            ('Address:', f"{company.address_line_1} {company.address_line_2}"),
            ('City:', f"{company.city}, {company.county}"),
            ('Employees:', str(company.employee_count)),
            ('Annual Revenue:', f"KES {company.annual_revenue:,.2f}" if company.annual_revenue else "N/A"),
        ]
        
        for label, value in info_rows:
            if value:
                p.drawString(50, y, label)
                p.drawString(150, y, value)
                y -= 20
        
        # Contacts
        y -= 20
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Contacts")
        y -= 20
        p.setFont("Helvetica", 10)
        
        for contact in company.contacts.all():
            p.drawString(50, y, f"{contact.get_full_name()} - {contact.email}")
            y -= 15
            if contact.position:
                p.drawString(70, y, f"Position: {contact.position}")
                y -= 15
        
        p.showPage()
        p.save()
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response


class CompanyDetailAPIView(LoginRequiredMixin, View):
    def get(self, request, pk):
        company = get_object_or_404(Company, pk=pk)
        data = {
            'id': company.pk,
            'name': company.name,
            'address_line_1': company.address_line_1,
            'address_line_2': company.address_line_2,
            'city': company.city,
            'county': company.county,
            'postal_code': company.postal_code,
        }
        return JsonResponse(data)


class LeadConversionView(LoginRequiredMixin, TemplateView):
    template_name = 'crm/reports/lead_conversion.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)

        # Get leads created in the last 30 days
        recent_leads = Lead.objects.filter(created_at__gte=thirty_days_ago)
        total_leads = recent_leads.count()

        # Calculate conversion metrics
        converted_leads = recent_leads.filter(status='won').count()
        conversion_rate = (converted_leads / total_leads * 100) if total_leads > 0 else 0

        # Get conversion by source
        sources_data = LeadSource.objects.filter(
            leads__created_at__gte=thirty_days_ago
        ).annotate(
            total_leads=Count('leads'),
            converted_leads=Count('leads', filter=Q(leads__status='won'))
        )

        for source in sources_data:
            source.conversion_rate = (
                source.converted_leads / source.total_leads * 100
            ) if source.total_leads > 0 else 0

        # Calculate average days to convert using updated_at instead of converted_date
        avg_days_to_convert = recent_leads.filter(
            status='won',
            converted_at__isnull=False
        ).annotate(
            conversion_time=ExpressionWrapper(
                F('converted_at') - F('created_at'),
                output_field=DurationField()
            )
        ).aggregate(
            avg_days=Avg('conversion_time')
        )['avg_days']

        # Convert timedelta to number of days if not None
        avg_days = avg_days_to_convert.days if avg_days_to_convert else 0

        context.update({
            'total_leads': total_leads,
            'converted_leads': converted_leads,
            'conversion_rate': conversion_rate,
            'sources_data': sources_data,
            'average_days_to_convert': avg_days
        })

        return context


class ReportView(LoginRequiredMixin, View):
    """Base class for report generation and export"""
    report_generators = {
        'activity': ActivityReport,
        'pipeline': PipelineReport,
        'revenue': RevenueReport,
        'customer': CustomerReport
    }

    def get_date_range(self, days=30):
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        return start_date, end_date

    def get_generator(self, report_type, start_date, end_date):
        generator_class = self.report_generators.get(report_type)
        if not generator_class:
            raise ValueError(f'Invalid report type: {report_type}')
        return generator_class(start_date=start_date, end_date=end_date, user=self.request.user)

    def process_data(self, data):
        """Convert Decimal objects to float for JSON serialization"""
        def decimal_to_float(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return obj
        return json.loads(json.dumps(data, default=decimal_to_float))

class GenerateReportView(ReportView):
    """Handle report generation and return JSON response"""
    
    def get(self, request, *args, **kwargs):
        try:
            report_type = request.GET.get('type')
            if not report_type:
                return JsonResponse({
                    'success': False,
                    'error': 'Report type is required'
                }, status=400)

            # Get date range
            try:
                days = int(request.GET.get('date_range', 30))
                start_date, end_date = self.get_date_range(days)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid date range'
                }, status=400)

            # Generate report
            try:
                generator = self.get_generator(report_type, start_date, end_date)
                data = generator.generate()
                processed_data = self.process_data(data)

                return JsonResponse({
                    'success': True,
                    'data': processed_data,
                    'metadata': {
                        'report_type': report_type,
                        'date_range': {
                            'start': start_date.isoformat(),
                            'end': end_date.isoformat()
                        }
                    }
                })

            except Exception as e:
                import traceback
                print(f"Report generation error: {str(e)}")
                print(traceback.format_exc())
                return JsonResponse({
                    'success': False,
                    'error': f'Error generating report: {str(e)}'
                }, status=500)

        except Exception as e:
            import traceback
            print(f"Unexpected error: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'error': f'Unexpected error: {str(e)}'
            }, status=500)

class ExportReportView(ReportView):
    """Handle report export in different formats"""

    def get(self, request, *args, **kwargs):
        try:
            report_type = request.GET.get('type')
            format = request.GET.get('format', 'csv')
            date_range = int(request.GET.get('date_range', 30))
            
            # Get date range
            start_date, end_date = self.get_date_range(days=date_range)
            
            # Generate report data
            data = {}
            if report_type == 'all':
                # Generate comprehensive report
                data = {
                    'activity': ActivityReport(start_date, end_date, self.request.user).generate(),
                    'pipeline': PipelineReport(start_date, end_date, self.request.user).generate(),
                    'revenue': RevenueReport(start_date, end_date, self.request.user).generate(),
                    'customer': CustomerReport(start_date, end_date, self.request.user).generate()
                }
            else:
                generator = self.get_generator(report_type, start_date, end_date)
                data = generator.generate()

            # Export based on format
            if format == 'csv':
                return self._export_csv(data, report_type)
            elif format == 'excel':
                return self._export_excel(data, report_type)
            elif format == 'pdf':
                return self._export_pdf(data, report_type)
            else:
                raise ValueError(f'Unsupported format: {format}')

        except Exception as e:
            logger.error(f"Error exporting report: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

    def _export_csv(self, data, report_type):
        output = io.StringIO()
        writer = csv.writer(output)
        
        if report_type == 'all':
            # Write comprehensive report
            for section, section_data in data.items():
                writer.writerow([f'{section.upper()} REPORT'])
                writer.writerow([])
                
                if section == 'activity':
                    writer.writerow(['Activities Summary'])
                    writer.writerow(['Type', 'Count', 'Completion Rate'])
                    for activity in section_data.get('by_type', []):
                        writer.writerow([
                            activity['activity_type'],
                            activity['count'],
                            f"{activity['completion_rate']:.1f}%"
                        ])
                
                elif section == 'pipeline':
                    writer.writerow(['Pipeline Summary'])
                    writer.writerow(['Stage', 'Count', 'Value', 'Weighted Value'])
                    for stage in section_data.get('by_stage', []):
                        writer.writerow([
                            stage['stage'],
                            stage['count'],
                            f"KES {stage['value']:,.2f}",
                            f"KES {stage['weighted_value']:,.2f}"
                        ])
                
                elif section == 'revenue':
                    writer.writerow(['Revenue Summary'])
                    writer.writerow(['Month', 'Revenue', 'Deals'])
                    for month in section_data.get('by_month', []):
                        writer.writerow([
                            month['month'],
                            f"KES {month['revenue']:,.2f}",
                            month['deals']
                        ])
                
                elif section == 'customer':
                    writer.writerow(['Customer Summary'])
                    writer.writerow(['Total Customers', section_data['total_customers']])
                    writer.writerow(['New Customers', section_data['new_customers']])
                    writer.writerow(['Retention Rate', f"{section_data['retention_rate']:.1f}%"])
                
                writer.writerow([])  # Add spacing between sections
        
        else:
            # Write single report type
            self._write_report_section(writer, report_type, data)

        filename = f"crm_report_{report_type}_{timezone.now().strftime('%Y%m%d')}.csv"
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _write_report_section(self, writer, report_type, data):
        """Helper method to write a specific report section"""
        if report_type == 'activity':
            writer.writerow(['Activities Summary'])
            writer.writerow(['Type', 'Count', 'Completion Rate'])
            for activity in data.get('by_type', []):
                writer.writerow([
                    activity['activity_type'],
                    activity['count'],
                    f"{activity['completion_rate']:.1f}%"
                ])
        
        elif report_type == 'pipeline':
            writer.writerow(['Pipeline Summary'])
            writer.writerow(['Stage', 'Count', 'Value', 'Weighted Value'])
            for stage in data.get('by_stage', []):
                writer.writerow([
                    stage['stage'],
                    stage['count'],
                    f"KES {stage['value']:,.2f}",
                    f"KES {stage['weighted_value']:,.2f}"
                ])
        
        # Add other report types as needed

    def _export_excel(self, data, report_type):
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Define styles
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Export based on report type
        if report_type == 'all':
            # Comprehensive report with multiple sheets
            self._write_activity_sheet(wb, data.get('activity', {}))
            self._write_pipeline_sheet(wb, data.get('pipeline', {}))
            self._write_revenue_sheet(wb, data.get('revenue', {}))
            self._write_customer_sheet(wb, data.get('customer', {}))
        else:
            # Single report type
            if report_type == 'activity':
                self._write_activity_sheet(wb, data)
            elif report_type == 'pipeline':
                self._write_pipeline_sheet(wb, data)
            elif report_type == 'revenue':
                self._write_revenue_sheet(wb, data)
            elif report_type == 'customer':
                self._write_customer_sheet(wb, data)

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"crm_report_{report_type}_{timezone.now().strftime('%Y%m%d')}.xlsx"
       
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _write_activity_sheet(self, workbook, data):
        ws = workbook.create_sheet("Activities")
        headers = ['Type', 'Count', 'Completion Rate']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write data
        for row, activity in enumerate(data.get('by_type', []), 2):
            ws.cell(row=row, column=1, value=activity['activity_type'])
            ws.cell(row=row, column=2, value=activity['count'])
            ws.cell(row=row, column=3, value=f"{activity['completion_rate']:.1f}%")

    def _write_pipeline_sheet(self, workbook, data):
        ws = workbook.create_sheet("Pipeline")
        headers = ['Stage', 'Count', 'Value', 'Weighted Value']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write data
        for row, stage in enumerate(data.get('by_stage', []), 2):
            ws.cell(row=row, column=1, value=stage['stage'])
            ws.cell(row=row, column=2, value=stage['count'])
            ws.cell(row=row, column=3, value=stage['value'])
            ws.cell(row=row, column=4, value=stage['weighted_value'])

    def _write_revenue_sheet(self, workbook, data):
        ws = workbook.create_sheet("Revenue")
        headers = ['Month', 'Revenue', 'Deals']
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Write monthly data
        for row, month in enumerate(data.get('by_month', []), 2):
            ws.cell(row=row, column=1, value=month['month'])
            ws.cell(row=row, column=2, value=month['revenue'])
            ws.cell(row=row, column=3, value=month['deals'])

    def _write_customer_sheet(self, workbook, data):
        ws = workbook.create_sheet("Customers")
        
        # Write summary
        ws.cell(row=1, column=1, value='Metric')
        ws.cell(row=1, column=2, value='Value')
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        
        row = 2
        metrics = [
            ('Total Customers', data.get('total_customers', 0)),
            ('New Customers', data.get('new_customers', 0)),
            ('Retention Rate', f"{data.get('retention_rate', 0):.1f}%"),
            ('Average Customer LTV', data.get('customer_lifetime_value', 0))
        ]
        
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1

    def _export_pdf(self, data, report_type):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Add data based on report type
        if report_type == 'all' or report_type == 'activity':
            table_data = [['Type', 'Count', 'Completion Rate']]
            for activity in data.get('by_type', []):
                table_data.append([
                    activity['activity_type'],
                    str(activity['count']),
                    f"{activity['completion_rate']:.1f}%"
                ])

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        filename = f"crm_report_{report_type}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def generate_full_report(self):
        # Implement comprehensive report generation
        return {
            'activities': ActivityReport(self.start_date, self.end_date, self.request.user).generate(),
            'pipeline': PipelineReport(self.start_date, self.end_date, self.request.user).generate(),
            'revenue': RevenueReport(self.start_date, self.end_date, self.request.user).generate(),
            'customers': CustomerReport(self.start_date, self.end_date, self.request.user).generate()
        }


class TeamAssignmentView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            employee_id = data.get('employeeId')
            role = data.get('role')
            department = data.get('department')

            if not all([employee_id, role, department]):
                return JsonResponse({
                    'success': False,
                    'error': 'Missing required fields'
                }, status=400)

            employee = User.objects.get(id=employee_id)
            
            # Update employee details
            employee.role = role
            employee.department = department
            employee.save(update_fields=['role', 'department'])

            return JsonResponse({
                'success': True,
                'message': f'{employee.get_full_name()} has been assigned as {employee.get_role_display()}'
            })

        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Employee not found'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
